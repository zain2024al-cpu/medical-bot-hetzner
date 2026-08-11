# services/residency_archive_excel.py
# تصدير أرشيف الإقامات إلى ملف Excel — المرضى ومرافقوهم مع تواريخ انتهاء
# الإقامة والجواز، مرتّبين حسب الأقرب انتهاءً.
#
# لماذا Excel لا PDF: مسار PDF في هذا المشروع يعتمد WeasyPrint (مكتبات نظام
# على السيرفر — تعطّلت فعلياً هذه الجلسة بنقص libpango)، بينما openpyxl
# بايثون خالص بلا اعتماديات نظام. وهو أنسب للمتابعة: فرز وتصفية وطباعة.
# نفس نمط RTL المُثبت في services/pharmacy_evacuation_excel.py.

from __future__ import annotations

import io
import logging
from datetime import date, datetime

logger = logging.getLogger(__name__)

# ألوان حالة الانتهاء — تجعل الملف قابلاً للمسح البصري كالصورة
_FILL_EXPIRED = "FFCDD2"   # أحمر فاتح — منتهية
_FILL_URGENT  = "FFE0B2"   # برتقالي فاتح — ≤ 30 يوم
_FILL_SOON    = "FFF9C4"   # أصفر فاتح — ≤ 90 يوم
_FILL_OK      = None       # بلا تلوين


def _parse(raw: str | None) -> date | None:
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(raw)[:10], fmt).date()
        except Exception:
            continue
    return None


def _fmt(raw: str | None) -> str:
    d = _parse(raw)
    return d.strftime("%d/%m/%Y") if d else "—"


def _days_text(raw: str | None) -> tuple[str, str | None]:
    """(نص المتبقّي, لون التعبئة) لتاريخ انتهاء."""
    d = _parse(raw)
    if d is None:
        return "—", None
    delta = (d - date.today()).days
    if delta < 0:
        return f"منتهية منذ {abs(delta)} يوم", _FILL_EXPIRED
    if delta == 0:
        return "تنتهي اليوم", _FILL_EXPIRED
    if delta <= 30:
        return f"{delta} يوم", _FILL_URGENT
    if delta <= 90:
        return f"{delta} يوم", _FILL_SOON
    return f"{delta} يوم", _FILL_OK


def collect_archive_rows() -> list[dict]:
    """
    صف واحد لكل شخص (مريض أو مرافق)، والمرافق يلي مريضه مباشرة.

    الترتيب: حسب أقرب تاريخ انتهاء إقامة للمجموعة (المريض ومرافقيه معاً)،
    فتظهر الحالات العاجلة أعلى الملف. من بلا تاريخ يُدفع للأسفل.
    """
    from sqlalchemy import or_
    from db.session import SessionLocal
    from db.models import ResidencyProfile, ResidencyCompanion
    from modules.residency.views import format_status

    rows: list[dict] = []
    with SessionLocal() as db:
        # ✅ الإضافة اليدوية مستبعَدة من الأرشيف المُصدَّر (Excel/PDF كلاهما
        # يستدعيان هذه الدالة) — نفس نطاق قائمة الإقامات داخل البوت.
        profiles = (
            db.query(ResidencyProfile)
            .filter(or_(ResidencyProfile.source != "manual", ResidencyProfile.source.is_(None)))
            .order_by(ResidencyProfile.id)
            .all()
        )
        groups = []
        for p in profiles:
            comps = (
                db.query(ResidencyCompanion)
                .filter(ResidencyCompanion.profile_id == p.id)
                .order_by(ResidencyCompanion.id)
                .all()
            )
            dates = [_parse(p.expiry_date)] + [_parse(c.expiry_date) for c in comps]
            dates = [d for d in dates if d]
            groups.append((min(dates) if dates else date.max, p, comps))

        groups.sort(key=lambda g: g[0])

        for idx, (_, p, comps) in enumerate(groups, start=1):
            rows.append({
                "seq":        idx,
                "name":       p.name or "—",
                "role":       "مريض",
                "is_comp":    False,
                "status":     format_status(p.status),
                "res_number": p.residency_number or "—",
                "res_expiry": p.expiry_date,
                "pas_expiry": p.passport_expiry,
            })
            for c in comps:
                rows.append({
                    "seq":        "",
                    "name":       c.name or "—",
                    "role":       "مرافق",
                    "is_comp":    True,
                    "status":     format_status(c.status),
                    "res_number": c.residency_number or "—",
                    "res_expiry": c.expiry_date,
                    "pas_expiry": c.passport_expiry,
                })
    return rows


def build_archive_excel(rows: list[dict]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "أرشيف الإقامات"
    ws.sheet_view.rightToLeft = True

    title_font  = Font(name="Arial", bold=True, size=18, color="424242")
    sub_font    = Font(name="Arial", size=10, color="616161")
    header_font = Font(name="Arial", bold=True, size=11, color="212121")
    header_fill = PatternFill("solid", start_color="F0F0F0", end_color="F0F0F0")
    name_font   = Font(name="Arial", bold=True, size=10)
    comp_font   = Font(name="Arial", size=10, color="555555")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right       = Alignment(horizontal="right",  vertical="center", wrap_text=True)
    thin = Border(*[Side(style="thin", color="CCCCCC")] * 4)

    headers = ["م", "الاسم", "الصفة", "الحالة", "رقم الإقامة",
               "انتهاء الإقامة", "المتبقّي", "انتهاء الجواز", "المتبقّي"]
    widths  = [5, 28, 9, 14, 16, 15, 16, 15, 16]
    last    = get_column_letter(len(headers))

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = "أرشيف الإقامات — المرضى والمرافقون"
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = f"تاريخ التصدير: {date.today().strftime('%d/%m/%Y')}   •   الإجمالي: {len(rows)} شخص"
    ws["A2"].font = sub_font
    ws["A2"].alignment = center

    hdr = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr, column=i, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center, thin
    ws.row_dimensions[hdr].height = 22

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = hdr + 1
    for row in rows:
        res_txt, res_fill = _days_text(row["res_expiry"])
        pas_txt, pas_fill = _days_text(row["pas_expiry"])
        # المرافق يُزاح باسمه ليظهر تابعاً لمريضه بصرياً
        display_name = f"    ↳ {row['name']}" if row["is_comp"] else row["name"]

        values = [
            row["seq"], display_name, row["role"], row["status"], row["res_number"],
            _fmt(row["res_expiry"]), res_txt, _fmt(row["pas_expiry"]), pas_txt,
        ]
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = thin
            c.alignment = right if i == 2 else center
            c.font = comp_font if row["is_comp"] else name_font
            if i in (5,):
                c.number_format = "@"      # رقم الإقامة نصّاً — لا يُحوَّل لرقم
        if res_fill:
            ws.cell(row=r, column=7).fill = PatternFill("solid", start_color=res_fill, end_color=res_fill)
        if pas_fill:
            ws.cell(row=r, column=9).fill = PatternFill("solid", start_color=pas_fill, end_color=pas_fill)
        r += 1

    # تجميد الرأس + فلتر تلقائي — جوهر "سهولة المتابعة"
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    ws.auto_filter.ref = f"A{hdr}:{last}{max(r - 1, hdr)}"

    # مفتاح الألوان
    r += 1
    ws.cell(row=r, column=2, value="مفتاح الألوان:").font = header_font
    for label, color in (
        ("منتهية", _FILL_EXPIRED), ("خلال 30 يوم", _FILL_URGENT), ("خلال 90 يوم", _FILL_SOON),
    ):
        r += 1
        c = ws.cell(row=r, column=2, value=label)
        c.fill = PatternFill("solid", start_color=color, end_color=color)
        c.alignment = right
        c.border = thin

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(f"[residency.archive_excel] built  rows={len(rows)}")
    return buf


def build_residency_archive_export() -> tuple[io.BytesIO | None, int]:
    """(الملف, عدد الأشخاص) — (None, 0) إن كان الأرشيف فارغاً."""
    rows = collect_archive_rows()
    if not rows:
        return None, 0
    return build_archive_excel(rows), len(rows)
