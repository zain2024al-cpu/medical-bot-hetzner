# services/residency_table_export.py
# جدول حالات الإقامة — PDF منسَّق أو Excel، لحالة واحدة أو للكل.
#
# ✅ نفس أسلوب services/pharmacy_evacuation_pdf.py المُثبت: reportlab مباشر
# + services/pdf_arabic.py لإعادة التشكيل والـbidi.
#
# ⚠️ **اتجاه الأعمدة**: reportlab لا يعكس ترتيب الأعمدة لمستندات RTL —
# `hAlign` يُحاذي الجدول ككتلة فقط. فتُعرَّف الأعمدة **معكوسة صراحةً**
# (الملفات أولاً… حتى "م" أخيراً) ليظهر "م" في أقصى اليمين كما يُقرأ.
#
# ⚠️ **عرض الأعمدة يُحسب من عرض المحتوى** لا بقيم ثابتة، وإلا تجاوز
# الجدولُ المساحةَ بين الهامشين فبدا ملتصقاً بالحافة.

from __future__ import annotations

import io
import logging
import os
from datetime import date, datetime

logger = logging.getLogger(__name__)

_HERE = os.path.dirname(os.path.abspath(__file__))
_FONTS_DIR = os.path.normpath(os.path.join(_HERE, "..", "assets", "fonts"))

# الخط المضمَّن أولاً: هو الوحيد الموجود فعلاً على الخادم (Linux). خطوط
# ويندوز للتطوير المحلي فقط، والسقوط لـHelvetica يعني مربعات لا عربية.
_FONT_CANDIDATES = [
    (os.path.join(_FONTS_DIR, "Arabic-Regular.ttf"), "RnTblAr"),
    ("C:\\Windows\\Fonts\\arial.ttf", "Arial"),
    ("C:\\Windows\\Fonts\\tahoma.ttf", "Tahoma"),
]
_FONT_BOLD_CANDIDATES = [
    (os.path.join(_FONTS_DIR, "Arabic-Bold.ttf"), "RnTblArBd"),
    ("C:\\Windows\\Fonts\\arialbd.ttf", "ArialBd"),
]

# العناوين بترتيب القراءة الطبيعي (يمين ← يسار)؛ تُعكَس عند بناء الـPDF.
_HEADERS = ["م", "الاسم", "الصفة", "الحالة", "التنبيه", "الانتهاء", "المتبقّي", "الملفات"]
_WIDTH_SHARES = [0.045, 0.30, 0.075, 0.13, 0.115, 0.115, 0.13, 0.09]


def _remaining_text(expiry) -> str:
    """نصّ عمود المتبقّي — عدد أيام، أو «متأخّر ن»، أو «—» بلا تاريخ."""
    from modules.residency.days import days_until, _plural
    left = days_until(expiry)
    if left is None:
        return "—"
    if left > 0:
        return f"{left} {_plural(left)}"
    if left == 0:
        return "ينتهي اليوم"
    return f"متأخّر {abs(left)} {_plural(abs(left))}"


# حدّ «قرب الانتهاء» — دون هذا العدد من الأيام يُعتبَر مستعجلاً.
_SOON_DAYS = 30


def _urgency(expiry) -> str:
    """`over` منتهية · `soon` قاربت · `ok` بعيدة · `none` بلا تاريخ.

    ⚠️ تُحسب **مرة واحدة** هنا لا في كل مولِّد: لو حسبها PDF وExcel كلٌّ
    بمعياره لاختلف لون الصفّ نفسه بين الملفين.
    """
    from modules.residency.days import days_until
    left = days_until(expiry)
    if left is None:
        return "none"
    if left < 0:
        return "over"
    return "soon" if left <= _SOON_DAYS else "ok"


def collect_rows(families, status_label_of) -> list[dict]:
    """يفرد العائلات صفوفاً: المريض ثم مرافقوه.

    ⚠️ الترقيم **بالعائلة لا بالشخص**: كل قوائم البوت تَعُدّ الطلبات لا
    الأفراد، فترقيمٌ متسلسل للأشخاص هنا يُظهِر رقماً لا يطابق أي شاشة.
    المرافق يأخذ «↳» ليُقرأ تابعاً لمن فوقه.
    """
    from modules.residency.repository import get_file_counts

    ids = [p.id for f in families for p in ([f.root] + list(f.companions))]
    counts = get_file_counts(ids) if ids else {}

    rows = []
    for i, fam in enumerate(families, start=1):
        members = [(fam.root, True)] + [(c, False) for c in fam.companions]
        for person, is_root in members:
            rows.append({
                "num": str(i) if is_root else "↳",
                "name": person.name or "—",
                "role": "مريض" if is_root else "مرافق",
                "status": status_label_of(person.status),
                "remind": person.reminder_date or "—",
                "expiry": person.expiry_date or "—",
                "left": _remaining_text(person.expiry_date),
                "files": str(counts.get(person.id, 0)),
                "is_root": is_root,
                "urgency": _urgency(person.expiry_date),
            })
    return rows


# ────────────────────────────── PDF ──────────────────────────────

def build_pdf(rows: list[dict], title: str) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    from services.pdf_arabic import ar, pick_font

    FN = pick_font(_FONT_CANDIDATES)
    FB = pick_font(_FONT_BOLD_CANDIDATES, fallback=FN)

    page = landscape(A4)          # الأسماء العربية طويلة — العرضي أنسب
    margin = 1.2 * cm
    content_w = page[0] - 2 * margin

    head = ParagraphStyle("h", fontName=FB, fontSize=9, alignment=TA_CENTER,
                          textColor=colors.HexColor("#1F3864"), leading=12)
    cell = ParagraphStyle("c", fontName=FN, fontSize=8.5, alignment=TA_CENTER, leading=11)
    cell_r = ParagraphStyle("cr", fontName=FN, fontSize=8.5, alignment=TA_RIGHT, leading=11)
    # ⚠️ اللون على **الفقرة** لا على TableStyle: الخلايا هنا Paragraph،
    # و`TEXTCOLOR` في TableStyle لا يؤثّر على نصّ داخل فقرة لها نمطها.
    red_hot = ParagraphStyle("rh", parent=cell, fontName=FB,
                             textColor=colors.HexColor("#B00020"))
    red_soft = ParagraphStyle("rs", parent=cell,
                              textColor=colors.HexColor("#C0392B"))
    gray_none = ParagraphStyle("gn", parent=cell,
                               textColor=colors.HexColor("#9E9E9E"))

    def urgent_style(u):
        return {"over": red_hot, "soon": red_soft, "none": gray_none}.get(u, cell)
    ttl = ParagraphStyle("t", fontName=FB, fontSize=15, alignment=TA_CENTER,
                         textColor=colors.HexColor("#1F3864"), leading=20)
    sub = ParagraphStyle("s", fontName=FN, fontSize=9, alignment=TA_CENTER,
                         textColor=colors.HexColor("#666666"), leading=13)

    def P(txt, st=cell):
        return Paragraph(ar(txt), st)

    # ⚠️ معكوسة صراحةً — "م" آخر القائمة ليظهر في أقصى اليمين
    col_w = [content_w * s for s in reversed(_WIDTH_SHARES)]
    header = [P(h, head) for h in reversed(_HEADERS)]

    data = [header]
    root_lines = []
    for r in rows:
        if r["is_root"]:
            root_lines.append(len(data))
        us = urgent_style(r["urgency"])
        # العمودان معاً: التاريخ يقول متى، والمتبقّي يقول كم بقي — تلوين
        # أحدهما دون الآخر يجعل نصف السطر يصرخ ونصفه صامت.
        data.append([
            P(r["files"]), P(r["left"], us), P(r["expiry"], us), P(r["remind"]),
            P(r["status"]), P(r["role"]), P(r["name"], cell_r), P(r["num"]),
        ])

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=page, leftMargin=margin, rightMargin=margin,
                            topMargin=margin, bottomMargin=1.4 * cm,
                            title=title, author="Residency")
    story = [
        Paragraph(ar(title), ttl),
        Spacer(1, 0.15 * cm),
        Paragraph(ar(f"عدد الطلبات: {len(root_lines)}  ·  الأشخاص: {len(rows)}"
                     f"  ·  تاريخ الطباعة: {date.today().isoformat()}"), sub),
        Spacer(1, 0.35 * cm),
    ]

    if not rows:
        story.append(Paragraph(ar("لا توجد حالات بهذه الحالة."), sub))
    else:
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#DCE6F1")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#9BA5B4")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
        # صفّ المريض مظلَّل — الحدّ البصري بين عائلة وأخرى، فيُقرأ من تحته
        # تابعاً له بلا عمود إضافي.
        for ln in root_lines:
            style.append(("BACKGROUND", (0, ln), (-1, ln), colors.HexColor("#F2F6FC")))
        # المنتهية: خلفية وردية على خانتَي الانتهاء والمتبقّي (الفهرسان ١ و٢
        # بعد عكس الأعمدة) — تُلمَح من بعيد بلا قراءة الأرقام.
        for i, r in enumerate(rows, start=1):
            if r["urgency"] == "over":
                style.append(("BACKGROUND", (1, i), (2, i), colors.HexColor("#FCE8E6")))

        t = Table(data, colWidths=col_w, hAlign="CENTER", repeatRows=1)
        t.setStyle(TableStyle(style))
        story.append(t)

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont(FN, 8)
        canvas.setFillColor(colors.HexColor("#888888"))
        canvas.drawRightString(page[0] - margin, 0.7 * cm, ar(f"صفحة {doc_.page}"))
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return buf


# ───────────────────────────── Excel ─────────────────────────────

def build_excel(rows: list[dict], title: str) -> io.BytesIO:
    """⚠️ Excel لا يحتاج إعادة تشكيل ولا bidi — يتولّى العربية بنفسه.
    تمرير نصّ مُعاد تشكيله إليه يُنتِج حروفاً مفكّكة غير قابلة للبحث.
    لذلك تُكتَب النصوص هنا **خاماً** بعكس مسار الـPDF تماماً.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "الإقامات"
    ws.sheet_view.rightToLeft = True          # ورقة عربية من اليمين لليسار

    thin = Side(style="thin", color="9BA5B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ncols = len(_HEADERS)

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"].font = Font(bold=True, size=14, color="1F3864")
    ws["A1"].alignment = center

    ws.append([f"تاريخ الطباعة: {date.today().isoformat()}  ·  الأشخاص: {len(rows)}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A2"].alignment = center
    ws["A2"].font = Font(size=9, color="666666")

    ws.append([])
    ws.append(_HEADERS)                       # Excel يعكس العرض بنفسه
    hdr_fill = PatternFill("solid", fgColor="DCE6F1")
    for c in ws[4]:
        c.font = Font(bold=True, color="1F3864")
        c.fill = hdr_fill
        c.alignment = center
        c.border = border

    root_fill = PatternFill("solid", fgColor="F2F6FC")
    over_fill = PatternFill("solid", fgColor="FCE8E6")
    # نفس درجات ألوان الـPDF حرفياً — الملفان يعرضان البيانات نفسها،
    # فاختلاف اللون بينهما يجعل القارئ يشكّ في أيّهما الصحيح.
    URGENT_FONT = {
        "over": Font(bold=True, color="B00020"),
        "soon": Font(color="C0392B"),
        "none": Font(color="9E9E9E"),
    }
    EXPIRY_COL, LEFT_COL = 5, 6          # فهارس صفرية: الانتهاء، المتبقّي

    for r in rows:
        ws.append([r["num"], r["name"], r["role"], r["status"],
                   r["remind"], r["expiry"], r["left"], r["files"]])
        uf = URGENT_FONT.get(r["urgency"])
        for idx, c in enumerate(ws[ws.max_row]):
            c.border = border
            c.alignment = right if idx == 1 else center
            if r["is_root"]:
                c.fill = root_fill
            if idx in (EXPIRY_COL, LEFT_COL):
                if uf is not None:
                    c.font = uf
                if r["urgency"] == "over":
                    c.fill = over_fill    # يتقدّم على تظليل صفّ المريض

    for i, share in enumerate(_WIDTH_SHARES, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, round(share * 105))

    ws.freeze_panes = "A5"                    # يبقى الرأس ظاهراً عند التمرير
    if rows:
        ws.auto_filter.ref = f"A4:{get_column_letter(ncols)}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_filename(status_label: str, ext: str) -> str:
    """اسم ملف آمن — يُزال ما تمنعه أنظمة الملفات وتليجرام."""
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    bad = set('\\/:*?"<>|')
    safe = "".join(ch for ch in status_label if ch not in bad).strip()
    safe = safe.replace(" ", "_") or "الكل"
    return f"الإقامات_{safe}_{stamp}.{ext}"
