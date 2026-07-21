# services/pharmacy_evacuation_excel.py
# Excel لـ"🖨️ طباعة مسير إخلاء الأدوية والمستلزمات الطبية".
# نفس نمط openpyxl + rightToLeft المُثبت في bot/handlers/admin/admin_evaluation.py.
# نفس ترتيب أعمدة PDF بالضبط: م | المبلغ | الاسم | رقم الفاتورة | بند الصرف | البيان | التاريخ
#
# ✅ ملاحظة مهمة حول تنسيق التواريخ: بعض تطبيقات الجداول (WPS، إكسل على
# الجوال، Google Sheets، LibreOffice) تُعيد اكتشاف أي نص يُشبه تاريخاً
# عند فتح الملف وتُعيد تفسيره كرقم/تاريخ حسب إعداداتها الخاصة عند
# الاستيراد — حتى لو كانت الخلية مُخزَّنة فعلياً كنص صرف (data_type='s')
# مع number_format='@' في ملف openpyxl (بعض المستوردين يتجاهلون هذا
# التلميح تماماً). لا حرف LRM بعد السنة فقط ولا فاصل FULLWIDTH كانا
# كافيَين لكل التطبيقات (تأكَّدنا فعلياً: بعضها يتجاهل نوع الفاصل تماماً
# ويبحث فقط عن أي 8 أرقام متتالية بعد حذف كل ما هو غير رقم، ثم يعيد
# عرضها بتنسيقه الافتراضي المضغوط "20260621"). الحل الأكثر ضماناً: تبعيث
# حرف LRM بين كل رقمين على حدة (وليس مرة واحدة فقط) — فلا يوجد رقمان
# متتاليان بلا حرف غير مرئي بينهما، فيستحيل على أي نمط \d{2,} أن يلتقط
# التاريخ كوحدة واحدة مهما كان الفاصل الظاهري.

import io
import logging
from datetime import date, datetime

logger = logging.getLogger(__name__)

_TEXT_FORMAT = "@"
_LRM = "‎"  # Left-to-Right Mark — غير مرئي
_FW_SLASH = "／"  # FULLWIDTH SOLIDUS — يبدو كـ"/" لكنه ليس حرف الفاصل الذي تبحث عنه المستوردات


def _scatter(segment: str) -> str:
    """يُدرِج LRM بين كل رقمين متتاليَين — يكسر أي محاولة التقاط تعتمد
    على تجميع خانات متتالية بعد تجاهل الفواصل الظاهرة."""
    return _LRM.join(segment)


def _safe_date_text(dt) -> str:
    """نص تاريخ لا يمكن لأي تطبيق جداول إعادة تفسيره كرقم/تاريخ فعلي —
    يبدو "2026/06/21" لكن كل رقمين مفصولان بـLRM غير مرئي، وكل فاصل
    ظاهر هو FULLWIDTH SOLIDUS وليس شرطة مائلة حقيقية."""
    if isinstance(dt, date):
        y, m, d = _scatter(f"{dt.year:04d}"), _scatter(f"{dt.month:02d}"), _scatter(f"{dt.day:02d}")
        return f"{y}{_FW_SLASH}{m}{_FW_SLASH}{d}"
    return str(dt or "")


def build_evacuation_excel(rows: list[dict], start_date: date, end_date: date) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "مسير الإخلاء"
    ws.sheet_view.rightToLeft = True

    # ✅ مخطَّط رمادي/بيج مطابق لنسخة PDF (كان هذا الملف أزرق بالكامل ولم
    # يُحدَّث في تعديل "أزرق → رمادي" السابق). صف الرأس وصف الإجمالي بيج فاتح
    # (F0F0F0 = نفس شريط سند الصرف) بنص داكن؛ العناوين/التذييل رمادي داكن (424242).
    header_font = Font(name="Arial", bold=True, color="212121", size=11)
    header_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    bismillah_font = Font(name="Arial", bold=True, size=22, color="424242")
    title_font = Font(name="Arial", bold=True, size=20, color="424242")
    band_font = Font(name="Arial", bold=True, size=11, color="424242")
    band_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    normal_font = Font(name="Arial", size=10)
    total_font = Font(name="Arial", bold=True, size=11, color="212121")
    total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    footer_font = Font(name="Arial", bold=True, size=10, color="424242")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    band_border = Border(
        left=Side(style="thin", color="B0BEC5"), right=Side(style="thin", color="B0BEC5"),
        top=Side(style="thin", color="B0BEC5"), bottom=Side(style="thin", color="B0BEC5"),
    )
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["م", "المبلغ", "الاسم", "رقم الفاتورة", "بند الصرف", "البيان", "التاريخ"]
    last_col = get_column_letter(len(headers))
    n = len(headers)
    col_widths = [6, 24, 22, 13.26953125, 7.1796875, 26, 14]

    # ── بسم الله الرحمن الرحيم — خط أكبر، ورُفعت لأعلى بمساحة أوسع ────────────
    ws.row_dimensions[1].height = 10  # هامش علوي فارغ فوق البسملة
    ws["A1"] = "؛"
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "بسم الله الرحمن الرحيم"
    ws["A2"].font = bismillah_font
    ws["A2"].alignment = center_align
    ws.row_dimensions[2].height = 30

    # فراغ فاصل واضح قبل العنوان
    ws.row_dimensions[3].height = 10

    # ── العنوان الرئيسي ───────────────────────────────────────────────────────
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = "مسير إخلاء الأدوية والمستلزمات الطبية"
    ws["A4"].font = title_font
    ws["A4"].alignment = center_align
    ws.row_dimensions[4].height = 26

    ws.row_dimensions[5].height = 8

    # ── الحقول الثلاثة — كل حقل في مجموعة أعمدة مستقلة (وليس نصاً واحداً
    # طويلاً قد يُقتَطع)، مع مساحة فارغة واضحة (خانة) بجانب كل تسمية،
    # وحقل التاريخ يحصل على قالب "20__/__/__" جاهز للتعبئة اليدوية ──────────
    band_row = 6
    ws.row_dimensions[band_row].height = 30.5
    # نفس منطق توزيع صف التوقيعات: العمود 1 يظهر أقصى اليمين تلقائياً
    # بفضل rightToLeft=True، فتُقرأ المجموعات بترتيبها الطبيعي من العمود 1.
    cols_per_group = n / 3
    group_ranges = []
    for i in range(3):
        start = int(round(i * cols_per_group)) + 1
        end = int(round((i + 1) * cols_per_group))
        group_ranges.append((start, max(start, end)))

    # ✅ "20__‎م" وليس "20__م": إكسل يطبّق خوارزمية bidi تلقائياً على كل نص RTL
    # (بعكس PDF حيث نرسم يدوياً بلا bidi إطلاقاً) — وجود حرف عربي قوي (م)
    # ملاصقاً مباشرة لرقمين جزئيين وشرطتين سفليتين قد يجعل bidi إكسل يُعيد
    # ترتيب موضع "م" بالنسبة لـ"20". فاصل LRM هنا يُثبّت اتجاه رقم السنة
    # صراحة ويمنع أي تفاعل مع الحرف الذي يليه.
    band_fields = [
        ("رقم سند الصرف:", "____________"),
        ("رقم القيد:", "____________"),
        ("تاريخ تسليم المسير:", f"20__{_LRM}م / __ / __"),
    ]
    for (label, blank), (c_start, c_end) in zip(band_fields, group_ranges):
        if c_end > c_start:
            ws.merge_cells(start_row=band_row, start_column=c_start, end_row=band_row, end_column=c_end)
        cell = ws.cell(row=band_row, column=c_start, value=f"{label}  {blank}")
        cell.font = band_font
        cell.alignment = center_align
        cell.border = band_border
        for col in range(c_start, c_end + 1):
            ws.cell(row=band_row, column=col).fill = band_fill
            if col != c_start:
                ws.cell(row=band_row, column=col).border = band_border

    ws.row_dimensions[7].height = 10

    header_row_idx = 10
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row_idx, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    total_amount = 0.0
    for i, r in enumerate(rows, start=1):
        row_idx = header_row_idx + i
        date_str = _safe_date_text(r["date"])
        values = [i, f'{r["amount"]:.2f}', r["name"], r["invoice_number"], r["expense_item"], r["statement"], date_str]
        total_amount += r["amount"]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = center_align  # ✅ كل الأعمدة موسَّطة لمظهر أفضل وأكثر اتساقاً
            if col == len(headers):  # عمود التاريخ — نصّ صرف دائماً، لا يُعاد تفسيره كرقم
                cell.number_format = _TEXT_FORMAT
            elif col == 2:  # عمود المبلغ — نفس تنسيق العملة من العينة المرجعية (بلا أثر مرئي لأن القيمة نص)
                cell.number_format = "[$₹-439]#,##0.00"

    # ✅ قيمة الإجمالي تظهر تحت عمود "المبلغ" (العمود 2) تحديداً فقط —
    # لا تمتد عبر الجدول كاملاً. عمود "م" (1) يُترك فارغاً بنفس التظليل،
    # والتسمية "إجمالي المبلغ" تمتد من "الاسم" حتى "التاريخ" (الأعمدة 3-7).
    total_row_idx = header_row_idx + len(rows) + 1

    empty_cell = ws.cell(row=total_row_idx, column=1, value="")
    empty_cell.fill = total_fill

    amount_cell = ws.cell(row=total_row_idx, column=2, value=f"{total_amount:,.2f}")
    amount_cell.font = total_font
    amount_cell.fill = total_fill
    amount_cell.alignment = center_align

    ws.merge_cells(f"C{total_row_idx}:{last_col}{total_row_idx}")
    label_cell = ws.cell(row=total_row_idx, column=3, value="إجمالي المبلغ")
    label_cell.font = total_font
    label_cell.fill = total_fill
    label_cell.alignment = center_align
    for col in range(4, len(headers) + 1):
        ws.cell(row=total_row_idx, column=col).fill = total_fill

    # ── صف تذييل التوقيعات (فارغ دائماً) ──────────────────────────────────────
    # ✅ العمود 1 (A) يظهر أقصى يمين الشاشة تلقائياً بفضل rightToLeft=True —
    # لذا تُوزَّع التسميات بترتيبها الطبيعي بدءاً من العمود 1 (وليس معكوسة
    # كما في PDF، حيث لا يوجد انعكاس تلقائي مماثل)، فيظهر "مستلم العهدة"
    # في أقصى اليمين تلقائياً بلا أي حساب عكسي.
    # ✅ التجميع هنا يعتمد على عرض الأعمدة الفعلي (col_widths) وليس عددها —
    # عمود "بند الصرف" (E) ضيّق جداً (7.18) مقارنة ببقية الأعمدة، فتجميع
    # متساوي العدد كان يترك "المسؤول المالي" وحيداً بالكامل في هذا العمود
    # الضيق فيختفي تماماً عند الطباعة (تأكَّدنا من هذا مباشرة على عينة
    # حقيقية مرفوعة من المستخدم). التجميع بالعرض التراكمي (أقرب عمود لكل
    # ربع من إجمالي العرض) يضمن أن كل توقيع يحصل على مساحة كافية دائماً.
    footer_row_idx = total_row_idx + 3
    footer_labels_rtl = ["مستلم العهدة", "المراجعة", "المسؤول المالي", "مسؤول العمليات"]
    cum = []
    running = 0.0
    for w in col_widths:
        running += w
        cum.append(running)
    quarter = cum[-1] / 4
    boundaries = []
    prev = 0
    for k in (1, 2, 3):
        target = k * quarter
        search_range = range(prev, n - (3 - k))
        best_idx = min(search_range, key=lambda idx: abs(cum[idx] - target))
        boundaries.append(best_idx)
        prev = best_idx + 1
    col_ranges = []
    start = 1
    for b in boundaries:
        col_ranges.append((start, b + 1))
        start = b + 2
    col_ranges.append((start, n))

    # ✅ خط تسليم متقطّع كصفّ مستقل بذاته (وليس سطر ثانٍ عبر "\n" داخل نفس
    # الخلية) — الاعتماد على wrap_text + ارتفاع صف يدوي لعرض سطرين داخل
    # خلية واحدة تبيَّن أنه غير موثوق عبر كل عارضات الجداول (بعضها يعرض
    # السطر الأول فقط ويقتطع الثاني). كل صف هنا سطر واحد بسيط بارتفاعه
    # الطبيعي، فلا يعتمد على أي حساب التفاف إطلاقاً.
    dash_counts = {"مستلم العهدة": 23, "المراجعة": 20, "المسؤول المالي": 24, "مسؤول العمليات": 24}
    for label, (c_start, c_end) in zip(footer_labels_rtl, col_ranges):
        ws.merge_cells(start_row=footer_row_idx, start_column=c_start, end_row=footer_row_idx, end_column=c_end)
        label_cell = ws.cell(row=footer_row_idx, column=c_start, value=label)
        label_cell.font = footer_font
        label_cell.alignment = center_align

        ws.merge_cells(start_row=footer_row_idx + 1, start_column=c_start, end_row=footer_row_idx + 1, end_column=c_end)
        dash_cell = ws.cell(row=footer_row_idx + 1, column=c_start, value="-" * dash_counts[label])
        dash_cell.font = footer_font
        dash_cell.alignment = center_align
    ws.row_dimensions[footer_row_idx].height = 18
    ws.row_dimensions[footer_row_idx + 1].height = 18

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info(f"[pharmacy_evacuation_excel] built  rows={len(rows)}  size={output.getbuffer().nbytes:,}")
    return output
