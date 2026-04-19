# -*- coding: utf-8 -*-
"""
استعادة/دمج تقارير من نسخة SQLite احتياطية أو من Excel/CSV/JSON.
يُنشئ صفوفاً جديدة (بدون الحفاظ على id القديم) لتجنب تعارض المفاتيح.
"""
from __future__ import annotations

import csv
import json
import logging
import os
import re
import sqlite3
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import DateTime, inspect
from sqlalchemy.sql import sqltypes

from db.models import Report
from db.session import SessionLocal
from services.translators_service import resolve_translator_for_report

logger = logging.getLogger(__name__)

# أعمدة جدول reports حسب النموذج الحالي
_REPORT_COLUMN_NAMES = {c.key for c in inspect(Report).mapper.column_attrs}


def _parse_dt(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        try:
            return datetime.fromtimestamp(val)
        except Exception:
            return None
    s = str(val).strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s[:26], fmt)
        except Exception:
            continue
    return None


def _row_effective_date(row: Dict[str, Any]) -> Optional[datetime]:
    rd = _parse_dt(row.get("report_date"))
    ca = _parse_dt(row.get("created_at"))
    if rd:
        return rd
    return ca


def _filter_row_in_range(row: Dict[str, Any], start: datetime, end: datetime) -> bool:
    """نطاق [start, end) — end حصرية."""
    eff = _row_effective_date(row)
    if not eff:
        return False
    return start <= eff < end


def _clean_row_for_insert(row: Dict[str, Any], clear_fks: bool) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for k, v in row.items():
        if k not in _REPORT_COLUMN_NAMES:
            continue
        if k == "id":
            continue
        if clear_fks and k in ("patient_id", "hospital_id", "department_id", "doctor_id"):
            out[k] = None
            continue
        col = Report.__table__.columns.get(k)
        if col is None:
            continue
        ctype = col.type
        if isinstance(ctype, DateTime):
            out[k] = _parse_dt(v) if v is not None else None
            continue
        if isinstance(ctype, sqltypes.Boolean):
            if v is None:
                out[k] = None
            elif isinstance(v, (int, float)):
                out[k] = bool(v)
            else:
                out[k] = bool(v)
            continue
        if isinstance(ctype, sqltypes.Integer):
            if v is None:
                out[k] = None
            else:
                try:
                    out[k] = int(v)
                except (TypeError, ValueError):
                    out[k] = None
            continue
        out[k] = v
    if not out.get("status"):
        out["status"] = "active"
    return out


def _read_rows_from_sqlite(path: str) -> Tuple[List[str], List[tuple]]:
    conn = sqlite3.connect(f"file:{path}?mode=ro", uri=True)
    try:
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='reports'")
        if not cur.fetchone():
            return [], []
        cur.execute("SELECT * FROM reports")
        cols = [d[0] for d in cur.description]
        rows = cur.fetchall()
        return cols, rows
    finally:
        conn.close()


def merge_reports_from_sqlite(
    sqlite_path: str,
    start: datetime,
    end: datetime,
    *,
    clear_fks: bool = True,
) -> Tuple[int, int, List[str]]:
    """
    يقرأ تقارير من ملف SQLite خارجي ويُدرجها في قاعدة البيانات الحالية.
    يُفلتر بالفترة حسب COALESCE(report_date, created_at) معقولاً من الصف كقاموس.

    Returns: (inserted, skipped_out_of_range, error_messages)
    """
    cols, raw_rows = _read_rows_from_sqlite(sqlite_path)
    if not cols:
        return 0, 0, ["لا يوجد جدول reports في الملف أو الملف تالف"]

    inserted = 0
    skipped = 0
    errors: List[str] = []

    dict_rows: List[Dict[str, Any]] = []
    for tup in raw_rows:
        d = dict(zip(cols, tup))
        if not _filter_row_in_range(d, start, end):
            skipped += 1
            continue
        dict_rows.append(d)

    for d in dict_rows:
        session = SessionLocal()
        try:
            kwargs = _clean_row_for_insert(d, clear_fks=clear_fks)
            session.add(Report(**kwargs))
            session.commit()
            inserted += 1
        except Exception as e:
            session.rollback()
            errors.append(str(e)[:200])
            logger.warning("reports_recovery sqlite row fail: %s", e, exc_info=True)
        finally:
            session.close()

    return inserted, skipped, errors


def merge_reports_from_json(
    json_path: str,
    start: datetime,
    end: datetime,
    *,
    clear_fks: bool = True,
) -> Tuple[int, int, List[str]]:
    """
    ملف JSON: مصفوفة من الكائنات، أو { "reports": [ ... ] }.
    كل عنصر يجب أن يحتوي حقولاً مطابقة لأسماء أعمدة Report.
    """
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "reports" in data:
        data = data["reports"]
    if not isinstance(data, list):
        return 0, 0, ["تنسيق JSON غير صالح: توقع مصفوفة أو مفتاح reports"]

    inserted = 0
    skipped = 0
    errors: List[str] = []

    to_insert: List[Dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            skipped += 1
            continue
        if not _filter_row_in_range(item, start, end):
            skipped += 1
            continue
        to_insert.append(item)

    for d in to_insert:
        session = SessionLocal()
        try:
            kwargs = _clean_row_for_insert(d, clear_fks=clear_fks)
            session.add(Report(**kwargs))
            session.commit()
            inserted += 1
        except Exception as e:
            session.rollback()
            errors.append(str(e)[:200])
            logger.warning("reports_recovery json row fail: %s", e, exc_info=True)
        finally:
            session.close()

    return inserted, skipped, errors


def safe_unlink(path: Optional[str]) -> None:
    if path and os.path.isfile(path):
        try:
            os.remove(path)
        except OSError:
            pass


# ─── استيراد Excel / CSV: صف عناوين ثم بيانات — يُستورد كل الصفوف ───

# (اسم الحقل في Report، قائمة مرادفات للعنوان في الملف)
_HEADER_ALIASES: List[Tuple[str, Tuple[str, ...]]] = [
    ("patient_name", ("patient_name", "اسم المريض", "المريض", "patient", "patient name")),
    ("report_date", ("report_date", "تاريخ التقرير", "التاريخ", "date", "report date")),
    ("created_at", ("created_at", "تاريخ الإنشاء", "created")),
    ("translator_id", ("translator_id", "معرف المترجم", "tg_user_id", "telegram_id")),
    ("translator_name", ("translator_name", "المترجم", "translator", "اسم المترجم")),
    ("medical_action", ("medical_action", "نوع الإجراء", "الإجراء الطبي", "الإجراء", "medical action", "action")),
    ("hospital_name", ("hospital_name", "المستشفى", "hospital", "hospital name")),
    ("department", ("department", "القسم", "dept", "department_name")),
    ("doctor_name", ("doctor_name", "الطبيب", "doctor", "doctor name")),
    ("complaint_text", ("complaint_text", "الشكوى", "complaint")),
    ("doctor_decision", ("doctor_decision", "قرار الطبيب", "decision")),
    ("diagnosis", ("diagnosis", "التشخيص")),
    ("treatment_plan", ("treatment_plan", "خطة العلاج")),
    ("notes", ("notes", "ملاحظات", "note")),
    ("case_status", ("case_status", "حالة المريض")),
    ("followup_date", ("followup_date", "موعد العودة")),
    ("followup_reason", ("followup_reason", "سبب العودة")),
    ("room_number", ("room_number", "رقم الغرفة", "الغرفة")),
    ("status", ("status", "الحالة")),
]


def _norm_header_cell(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).replace("\ufeff", "").strip()
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def _resolve_header_to_field(header_cell: Any) -> Optional[str]:
    raw = str(header_cell).replace("\ufeff", "").strip() if header_cell is not None else ""
    if not raw:
        return None
    n = _norm_header_cell(raw)
    for field, aliases in _HEADER_ALIASES:
        for a in aliases:
            if a.lower() == n or _norm_header_cell(a) == n:
                return field
            if raw == a or raw.lower() == a.lower():
                return field
    # مطابقة snake_case إنجليزي
    snake = re.sub(r"\s+", "_", n)
    if snake in _REPORT_COLUMN_NAMES:
        return snake
    return None


def _insert_one_report_dict(d: Dict[str, Any], clear_fks: bool) -> None:
    session = SessionLocal()
    try:
        kwargs = _clean_row_for_insert(d, clear_fks=clear_fks)
        if kwargs.get("translator_name"):
            tid, canon = resolve_translator_for_report(session, str(kwargs["translator_name"]))
            if kwargs.get("translator_id") is None and tid is not None:
                kwargs["translator_id"] = tid
            kwargs["translator_name"] = canon
        if kwargs.get("report_date") is None and kwargs.get("created_at"):
            kwargs["report_date"] = kwargs["created_at"]
        if kwargs.get("report_date") is None:
            raise ValueError("لا يوجد تاريخ (report_date أو created_at)")
        session.add(Report(**kwargs))
        session.commit()
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


def insert_report_dicts_all(rows: List[Dict[str, Any]], *, clear_fks: bool = True) -> Tuple[int, int, List[str]]:
    """إدراج قائمة صفوف؛ يتخطى الصفوف الفارغة أو بدون تاريخ."""
    inserted = 0
    skipped = 0
    errors: List[str] = []
    for d in rows:
        if not d or not isinstance(d, dict):
            skipped += 1
            continue
        try:
            if d.get("report_date") is None and d.get("created_at") is None:
                skipped += 1
                continue
            _insert_one_report_dict(d, clear_fks)
            inserted += 1
        except Exception as e:
            errors.append(str(e)[:220])
            logger.warning("insert_report_dicts_all: %s", e, exc_info=True)
    return inserted, skipped, errors


def merge_reports_from_excel(excel_path: str, *, clear_fks: bool = True) -> Tuple[int, int, List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0, 0, ["مكتبة openpyxl غير متوفرة"]

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return 0, 0, ["الملف فارغ"]

        field_by_col: List[Optional[str]] = []
        for cell in header_row:
            field_by_col.append(_resolve_header_to_field(cell))

        if not any(f == "report_date" or f == "created_at" for f in field_by_col if f):
            return 0, 0, [
                "الصف الأول يجب أن يحتوي عمود **تاريخ التقرير** أو **report_date** "
                "(أو created_at). راجع العناوين."
            ]

        dict_rows: List[Dict[str, Any]] = []
        for row in rows_iter:
            if row is None:
                continue
            d: Dict[str, Any] = {}
            for i, val in enumerate(row):
                if i >= len(field_by_col):
                    break
                fn = field_by_col[i]
                if not fn:
                    continue
                if val is not None and str(val).strip() != "":
                    d[fn] = val
            if d:
                dict_rows.append(d)

        return insert_report_dicts_all(dict_rows, clear_fks=clear_fks)
    finally:
        wb.close()


def merge_reports_from_csv(csv_path: str, *, clear_fks: bool = True) -> Tuple[int, int, List[str]]:
    """يقرأ CSV بترميز UTF-8 أو cp1256 تلقائياً."""
    last_err: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "cp1256"):
        dict_rows: List[Dict[str, Any]] = []
        try:
            with open(csv_path, "r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except Exception:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                header_row = next(reader, None)
                if not header_row:
                    return 0, 0, ["ملف CSV فارغ"]

                field_by_col = [_resolve_header_to_field(h) for h in header_row]
                if not any(f == "report_date" or f == "created_at" for f in field_by_col if f):
                    return 0, 0, ["عناوين CSV: أضف عمود تاريخ التقرير أو report_date"]

                for row in reader:
                    if not row or not any(str(c).strip() for c in row):
                        continue
                    d: Dict[str, Any] = {}
                    for i, val in enumerate(row):
                        if i >= len(field_by_col):
                            break
                        fn = field_by_col[i]
                        if not fn:
                            continue
                        if val is not None and str(val).strip() != "":
                            d[fn] = val
                    if d:
                        dict_rows.append(d)
            return insert_report_dicts_all(dict_rows, clear_fks=clear_fks)
        except UnicodeDecodeError as e:
            last_err = e
            continue
    return 0, 0, [f"تعذر قراءة CSV: {last_err!s}" if last_err else "تعذر قراءة CSV"]


def merge_reports_from_json_full(json_path: str, *, clear_fks: bool = True) -> Tuple[int, int, List[str]]:
    """استيراد كل عناصر JSON دون تصفية بتاريخ (مناسب لأرشيف يناير/فبراير المحضّر يدوياً)."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "reports" in data:
        data = data["reports"]
    if not isinstance(data, list):
        return 0, 0, ["JSON: توقع مصفوفة [...] أو { \"reports\": [...] }"]

    dict_rows: List[Dict[str, Any]] = []
    for item in data:
        if isinstance(item, dict):
            dict_rows.append(item)
    return insert_report_dicts_all(dict_rows, clear_fks=clear_fks)
