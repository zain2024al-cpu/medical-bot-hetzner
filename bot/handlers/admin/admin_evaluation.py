# ================================================
# bot/handlers/admin/admin_evaluation.py
# نظام تقييم المترجمين - يعتمد على services/stats_service.py فقط
# ================================================

import logging
import io
import os
import re
import sys
import calendar
from datetime import datetime, date, timedelta
from sqlalchemy import text
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes, ConversationHandler, MessageHandler,
    CallbackQueryHandler, CommandHandler, filters
)
from telegram.constants import ParseMode
from db.session import SessionLocal
from db.models import MonthlyEvaluation, TranslatorDirectory
from bot.shared_auth import is_admin
from services.stats_service import get_monthly_stats, get_translator_stats, ALL_ACTION_TYPES
from services.inline_calendar import MONTHS_AR, DAYS_AR, format_date_arabic

logger = logging.getLogger(__name__)

# ════════════════════════════════════════
# حالات المحادثة
# ════════════════════════════════════════
(
    EVAL_SELECT_MODE,
    EVAL_SELECT_TRANSLATOR,
    EVAL_SELECT_YEAR,
    EVAL_SELECT_MONTH,
    EVAL_SELECT_PERIOD,
    EVAL_SELECT_DAY,
    EVAL_SELECT_FORMAT,
    EVAL_CUSTOM_START,
    EVAL_CUSTOM_END,
) = range(9)

EVAL_TRANSLATORS_PER_PAGE = 12

MONTH_NAMES = {
    1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل",
    5: "مايو", 6: "يونيو", 7: "يوليو", 8: "أغسطس",
    9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر"
}


# ════════════════════════════════════════
# أدوات مساعدة
# ════════════════════════════════════════

def _rating_label(percentage):
    """4 مستويات: ممتاز - جيد - مقبول - ضعيف"""
    if percentage >= 80:
        return "ممتاز", "🟢", "⭐⭐⭐⭐"
    elif percentage >= 60:
        return "جيد", "🟡", "⭐⭐⭐"
    elif percentage >= 40:
        return "مقبول", "🟠", "⭐⭐"
    else:
        return "ضعيف", "🔴", "⭐"


def _medal(rank):
    if rank == 1: return "🥇"
    elif rank == 2: return "🥈"
    elif rank == 3: return "🥉"
    return f"#{rank}"


def _report_label(count: int) -> str:
    return "تقرير" if count == 1 else "تقارير"


def _normalize_person_name(name: str) -> str:
    """تطبيع بسيط لأسماء الأشخاص لتقليل مشاكل اختلاف الهمزات/المسافات."""
    s = (name or "").strip().lower()
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا").replace("ٱ", "ا")
    s = s.replace("ة", "ه").replace("ى", "ي")
    return s


def _get_daily_counts(session, translator_id, start_date, end_date, translator_name=None):
    """جلب عدد التقارير اليومية - بالاسم أولاً (لتوحيد المترجمين) ثم بالـ ID"""
    if not translator_id and not translator_name:
        return []
    # شرط التاريخ المزدوج (يلتقط التقارير القديمة بـ UTC)
    _DF = """(
        (COALESCE(r.report_date, r.created_at) >= :start AND COALESCE(r.report_date, r.created_at) < :end)
        OR (DATE(datetime(r.created_at, '+5 hours', '+30 minutes')) >= :start AND DATE(datetime(r.created_at, '+5 hours', '+30 minutes')) < :end)
    )"""
    if translator_name:
        sql = text(f"""
            SELECT DATE(COALESCE(r.report_date, r.created_at)) as day, COUNT(*) as count
            FROM reports r
            LEFT JOIN translators td ON r.translator_id = td.translator_id
            WHERE {_DF}
            AND r.status = 'active'
            AND COALESCE(td.name, r.translator_name) = :tname
            GROUP BY day
            ORDER BY day
        """)
        rows = session.execute(sql, {"start": start_date, "end": end_date, "tname": translator_name}).fetchall()
    else:
        sql = text(f"""
            SELECT DATE(COALESCE(r.report_date, r.created_at)) as day, COUNT(*) as count
            FROM reports r
            WHERE {_DF}
            AND r.status = 'active'
            AND r.translator_id = :translator_id
            GROUP BY day
            ORDER BY day
        """)
        rows = session.execute(sql, {"start": start_date, "end": end_date, "translator_id": translator_id}).fetchall()
    return [(str(r[0]), int(r[1])) for r in rows]


async def _send_text_chunks(message, text):
    lines = text.splitlines()
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > 3500:
            try:
                await message.reply_text(chunk, parse_mode=ParseMode.MARKDOWN)
            except Exception:
                await message.reply_text(chunk)
            chunk = ""
        chunk += (line + "\n")
    if chunk.strip():
        try:
            await message.reply_text(chunk.strip(), parse_mode=ParseMode.MARKDOWN)
        except Exception:
            await message.reply_text(chunk.strip())


def _compute_rating(stats_results):
    """
    يحسب عدد التقارير قبل 8 مساءً (= إجمالي - متأخر) لكل مترجم، ويُرتّب النتائج
    من الأعلى إجمالي تقارير إلى الأدنى، ثم الأقل تأخراً عند التعادل.

    ملاحظة: بناءً على طلب الإدارة تمت إزالة حقول التقييم (النسبة/المستوى/النجوم)
    من العرض، لكن نُبقي حقول التوافق القديمة بقيم صفرية لكيلا تنكسر
    الاستدعاءات الأخرى التي تحفظ في قاعدة البيانات.
    """
    if not stats_results:
        return []

    results = []
    for s in stats_results:
        total = int(s.get('total_reports') or 0)
        late = int(s.get('late_reports') or 0)
        before_8pm = max(total - late, 0)

        results.append({
            **s,
            'before_8pm_reports': before_8pm,
            # للتوافق مع الحفظ القديم فقط — ليست معروضة للمستخدم
            'final_score': 0,
            'punctuality_pct': 0,
            'level': '-',
            'color': '',
            'stars': '',
        })

    results.sort(key=lambda x: (-x['total_reports'], x['late_reports']))
    return results


# ════════════════════════════════════════
# توليد ملف PDF (reportlab على Windows)
# ════════════════════════════════════════

def _generate_pdf(results, period_label, year, month, start_date_str=None, end_date_str=None, target_name: str | None = None):
    """توليد تقرير PDF - بطاقة لكل مترجم"""

    total_reports = sum(r['total_reports'] for r in results)
    total_late = sum(r['late_reports'] for r in results)

    if not start_date_str or not end_date_str:
        if month == "all" or month == 0:
            start_date_str = f"01/01/{year}"
            end_date_str = f"31/12/{year}"
        else:
            m = int(month)
            start_date_str = f"01/{m:02d}/{year}"
            if m == 12:
                end_date_str = f"31/12/{year}"
            else:
                last_day = (datetime(year, m + 1, 1) - timedelta(days=1)).day
                end_date_str = f"{last_day}/{m:02d}/{year}"

    # ─── محاولة استخدام reportlab ───
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from arabic_reshaper import reshape
        from bidi.algorithm import get_display
    except Exception as e:
        logger.warning(f"فشل تحميل مكتبات PDF: {e}")
        return _generate_html_fallback(results, period_label, year, month, start_date_str, end_date_str, total_reports, total_late, target_name=target_name), "html"

    # تسجيل خط عربي
    font_name = "Helvetica"
    font_options = [
        ("C:\\Windows\\Fonts\\cairo.ttf", "Cairo"),
        ("C:\\Windows\\Fonts\\Cairo-Regular.ttf", "Cairo"),
        ("C:\\Windows\\Fonts\\tajawal.ttf", "Tajawal"),
        ("C:\\Windows\\Fonts\\Tajawal-Regular.ttf", "Tajawal"),
        ("C:\\Windows\\Fonts\\tahoma.ttf", "Tahoma"),
        ("C:\\Windows\\Fonts\\arial.ttf", "Arial"),
        ("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", "DejaVuSans"),
    ]
    for font_path, font_alias in font_options:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(font_alias, font_path))
                font_name = font_alias
                break
            except Exception:
                continue

    def r(text_val):
        value = "" if text_val is None else str(text_val)
        return get_display(reshape(value))

    # تصميم احترافي متعدد الصفحات
    from reportlab.platypus import Flowable
    from reportlab.graphics.shapes import Drawing, Rect, String, Line
    from reportlab.graphics.charts.barcharts import VerticalBarChart, HorizontalBarChart
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.lib.units import mm
    from services.stats_service import normalize_action_name as _norm_action

    MAIN = colors.HexColor("#0D47A1")
    MAIN_LIGHT = colors.HexColor("#E3F2FD")
    ACCENT = colors.HexColor("#26A69A")
    CARD_BG = colors.HexColor("#F8FAFC")
    GRID = colors.HexColor("#DDE3EA")
    # ── ألوان إضافية للعرض الإداري (لا تغيّر أي لون قائم) ──
    GREEN = colors.HexColor("#2E7D32")
    RED = colors.HexColor("#D32F2F")
    AMBER = colors.HexColor("#E65100")
    PURPLE = colors.HexColor("#7E57C2")
    MUTED = colors.HexColor("#607D8B")
    INK = colors.HexColor("#263238")

    class HeaderBand(Flowable):
        def __init__(self, title_text: str, subtitle_text: str):
            super().__init__()
            self.width = 540
            self.height = 62
            self.title_text = title_text
            self.subtitle_text = subtitle_text

        def draw(self):
            c = self.canv
            c.setFillColor(MAIN)
            c.roundRect(0, 0, self.width, self.height, 8, stroke=0, fill=1)
            # شعار مبسط (badge) بدلاً من صورة
            c.setFillColor(colors.white)
            c.circle(24, self.height / 2, 14, stroke=0, fill=1)
            c.setFillColor(MAIN)
            c.setFont(font_name, 14)
            c.drawCentredString(24, self.height / 2 - 5, "T")
            c.setFillColor(colors.white)
            c.setFont(font_name, 24)
            c.drawRightString(self.width - 12, self.height - 34, self.title_text)
            if self.subtitle_text:
                c.setFont(font_name, 11)
                c.setFillColor(colors.HexColor("#D6E4FF"))
                c.drawRightString(self.width - 12, 12, self.subtitle_text)

    def _draw_card(label: str, value: str, width=250, height=84, color=MAIN, value_color=None, label_color=None):
        d = Drawing(width, height)
        d.add(Rect(0, 0, width, height, rx=8, ry=8, fillColor=CARD_BG, strokeColor=GRID, strokeWidth=0.8))
        d.add(Rect(0, height - 8, width, 8, fillColor=color, strokeColor=color, strokeWidth=0))
        d.add(String(
            width - 10,
            height - 32,
            r(label),
            fontName=font_name,
            fontSize=12,
            fillColor=(label_color or colors.HexColor("#455A64")),
            textAnchor="end",
        ))
        d.add(String(
            width - 10,
            18,
            str(value),
            fontName=font_name,
            fontSize=24,
            fillColor=(value_color or MAIN),
            textAnchor="end",
        ))
        return d

    # ══════════════════════════════════════════════════════════════════════
    # أدوات عرض فقط (Presentation-only) — لا تُغيّر أي معادلة أو استعلام.
    # كل ما تفعله: قراءة الحقول الموجودة أصلاً في results وتجميعها للعرض.
    # ══════════════════════════════════════════════════════════════════════

    def _pct(part, whole) -> float:
        try:
            return (float(part) / float(whole) * 100.0) if float(whole) else 0.0
        except Exception:
            return 0.0

    def _i(it, key) -> int:
        try:
            return int(it.get(key) or 0)
        except Exception:
            return 0

    def _before8(it) -> int:
        """نفس التعبير المستخدم أصلاً في صفحة المترجم (بلا تغيير)."""
        return int(it.get("before_8pm_reports", max(_i(it, "total_reports") - _i(it, "late_reports"), 0)))

    def _late_pct(it) -> float:
        return _pct(_i(it, "late_reports"), _i(it, "total_reports"))

    def _norm_breakdown(it) -> dict:
        bd: dict[str, int] = {}
        for k, v in (it.get("action_breakdown") or {}).items():
            nk = _norm_action(k)
            if not nk:
                continue
            bd[nk] = bd.get(nk, 0) + int(v or 0)
        return {k: v for k, v in bd.items() if v > 0}

    def _top_action(it):
        bd = _norm_breakdown(it)
        if not bd:
            return "—", 0
        name, cnt = max(bd.items(), key=lambda x: x[1])
        return name, cnt

    def _avg_daily(it) -> float:
        wd = _i(it, "work_days")
        return (_i(it, "total_reports") / wd) if wd > 0 else 0.0

    def _kpi(label, value, color=MAIN, width=170, height=62, value_color=None):
        """بطاقة KPI مدمجة للصفحات الإدارية."""
        d = Drawing(width, height)
        d.add(Rect(0, 0, width, height, rx=6, ry=6, fillColor=CARD_BG, strokeColor=GRID, strokeWidth=0.7))
        d.add(Rect(0, height - 5, width, 5, fillColor=color, strokeColor=color, strokeWidth=0))
        d.add(String(width - 9, height - 23, r(label), fontName=font_name, fontSize=9.5,
                     fillColor=MUTED, textAnchor="end"))
        d.add(String(width - 9, 13, str(value), fontName=font_name, fontSize=19,
                     fillColor=(value_color or color), textAnchor="end"))
        return d

    def _kpi_grid(cells, per_row=3, cw=176):
        """شبكة KPI مرتبة RTL: أول عنصر يظهر أقصى اليمين."""
        rows = []
        for s in range(0, len(cells), per_row):
            chunk = cells[s:s + per_row]
            while len(chunk) < per_row:
                chunk.append(Drawing(170, 62))
            rows.append(list(reversed(chunk)))  # عكس ليقرأ من اليمين
        t = Table(rows, colWidths=[cw] * per_row)
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        return t

    def _pie(data, labels, palette, width=262, height=172, title=None):
        """رسم دائري آمن — يُعيد None إذا كانت كل القيم صفراً.
        الدائرة يساراً ووسيلة الإيضاح يميناً (قراءة عربية) بلا تداخل."""
        vals = [max(0, int(v or 0)) for v in data]
        if sum(vals) <= 0:
            return None
        d = Drawing(width, height)
        if title:
            d.add(String(width - 8, height - 13, r(title), fontName=font_name,
                         fontSize=10, fillColor=INK, textAnchor="end"))
        pie = Pie()
        pie.x = 8
        pie.y = 16
        pie.width = 96
        pie.height = 96
        pie.data = vals
        total_v = sum(vals)
        pie.labels = [f"{_pct(v, total_v):.0f}%" if v else "" for v in vals]
        pie.slices.strokeWidth = 0.6
        pie.slices.strokeColor = colors.white
        pie.slices.fontName = font_name
        pie.slices.fontSize = 8
        pie.sideLabels = False
        for idx, col in enumerate(palette[:len(vals)]):
            pie.slices[idx].fillColor = col
        d.add(pie)
        # وسيلة إيضاح على أقصى اليمين — بعيدة تماماً عن الدائرة
        ly = height - 38
        for idx, lab in enumerate(labels[:len(vals)]):
            d.add(Rect(width - 16, ly - 2, 9, 9, fillColor=palette[idx], strokeColor=palette[idx]))
            d.add(String(width - 21, ly, r(f"{lab} ({vals[idx]})"), fontName=font_name,
                         fontSize=8.5, fillColor=INK, textAnchor="end"))
            ly -= 16
        return d

    def _hbar(labels, values, color=MAIN, width=540, height=None, title=None):
        """رسم أعمدة أفقية — مناسب للأسماء العربية الطويلة.
        ✅ HorizontalBarChart يرسم أول عنصر في الأسفل، لذا نعكس القوائم حتى
        يظهر الأعلى قيمةً في الأعلى (ترتيب منطقي للقراءة الإدارية)."""
        vals = [max(0, int(v or 0)) for v in values]
        if not vals or sum(vals) <= 0:
            return None
        labels = list(labels)[::-1]
        vals = vals[::-1]
        n = len(vals)
        height = height or (28 + n * 19)
        d = Drawing(width, height)
        if title:
            d.add(String(width - 6, height - 13, r(title), fontName=font_name,
                         fontSize=10.5, fillColor=INK, textAnchor="end"))
        ch = HorizontalBarChart()
        ch.x = 34
        ch.y = 12
        ch.width = width - 150
        ch.height = height - 32
        ch.data = [vals]
        ch.strokeColor = GRID
        ch.valueAxis.valueMin = 0
        ch.valueAxis.valueMax = max(vals) + 1
        ch.valueAxis.valueStep = max(1, int((max(vals) + 1) / 4))
        ch.valueAxis.labels.fontName = font_name
        ch.valueAxis.labels.fontSize = 7.5
        ch.categoryAxis.categoryNames = [r(x) for x in labels]
        ch.categoryAxis.labels.fontName = font_name
        ch.categoryAxis.labels.fontSize = 8
        ch.categoryAxis.labels.dx = -4
        ch.bars[0].fillColor = color
        ch.bars[0].strokeColor = color
        ch.barWidth = 9
        d.add(ch)
        return d

    def _bullets(title, items, color, width=268):
        """صندوق نقاط (قوة / تحتاج تحسين) — نص مشتق من البيانات فقط."""
        items = items or ["—"]
        rows = [[Paragraph(r(title), ParagraphStyle(
            "bt", parent=styles["Normal"], fontName=font_name, fontSize=11,
            leading=15, alignment=TA_RIGHT, textColor=colors.white))]]
        for it_txt in items:
            rows.append([Paragraph(r(f"• {it_txt}"), ParagraphStyle(
                "bi", parent=styles["Normal"], fontName=font_name, fontSize=9.5,
                leading=14, alignment=TA_RIGHT, textColor=INK))])
        t = Table(rows, colWidths=[width])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), color),
            ("BACKGROUND", (0, 1), (0, -1), CARD_BG),
            ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 7),
            ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ]))
        return t

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=18, leftMargin=18, topMargin=14, bottomMargin=14)
    styles = getSampleStyleSheet()
    report_title_style = ParagraphStyle("report_title", parent=styles["Heading1"], fontName=font_name, fontSize=28, leading=34, alignment=TA_CENTER)
    date_style = ParagraphStyle("date", parent=styles["Normal"], fontName=font_name, fontSize=14, leading=18, alignment=TA_CENTER, textColor=colors.HexColor("#455A64"))
    base_style = ParagraphStyle("base", parent=styles["Normal"], fontName=font_name, fontSize=12, leading=16, alignment=TA_RIGHT)
    section_style = ParagraphStyle("section", parent=styles["Heading3"], fontName=font_name, fontSize=18, leading=24, alignment=TA_CENTER, textColor=MAIN)

    story = []
    total_before = total_reports - total_late

    # ══════════════════════════════════════════════════════════════════════
    # تجميعات للعرض فقط (مشتقة من نفس الحقول الموجودة في results)
    # ══════════════════════════════════════════════════════════════════════
    try:
        _sd_g = datetime.strptime(start_date_str, "%d/%m/%Y").date()
        _ed_g = datetime.strptime(end_date_str, "%d/%m/%Y").date()
        period_days_g = max(int((_ed_g - _sd_g).days) + 1, 1)
    except Exception:
        period_days_g = "-"

    n_tr = len(results)
    tot_before8 = sum(_before8(it) for it in results)
    tot_work_days = sum(_i(it, "work_days") for it in results)
    tot_yes = sum(_i(it, "paper_yes") for it in results)
    tot_no = sum(_i(it, "paper_no") for it in results)
    tot_pending = sum(_i(it, "paper_pending") for it in results)
    gate_total = tot_yes + tot_no + tot_pending
    late_pct_g = _pct(total_late, total_reports)
    upload_pct_g = _pct(tot_yes, gate_total)
    team_avg_daily = (total_reports / tot_work_days) if tot_work_days else 0.0

    # ترتيب العرض: كما يصل من stats_service (الأعلى تقارير أولاً)
    ranked = list(results)
    with_reports = [it for it in ranked if _i(it, "total_reports") > 0]

    def _best(key_fn, pool=None, reverse=True):
        pool = pool if pool is not None else with_reports
        if not pool:
            return None
        return sorted(pool, key=key_fn, reverse=reverse)[0]

    best_volume   = _best(lambda it: _i(it, "total_reports"))
    best_ontime   = _best(lambda it: (-_late_pct(it), _i(it, "total_reports")))
    best_upload   = _best(lambda it: _i(it, "paper_yes"))
    worst_late    = _best(lambda it: (_late_pct(it), _i(it, "total_reports")))
    worst_no      = _best(lambda it: _i(it, "paper_no"), pool=ranked)
    worst_pending = _best(lambda it: _i(it, "paper_pending"), pool=ranked)

    avg_no = (tot_no / n_tr) if n_tr else 0.0
    avg_pending = (tot_pending / n_tr) if n_tr else 0.0
    LATE_THRESHOLD = 30.0

    def _needs_followup(it):
        reasons = []
        if _i(it, "total_reports") > 0 and _late_pct(it) >= LATE_THRESHOLD:
            reasons.append(f"نسبة تأخير {_late_pct(it):.0f}%")
        if _i(it, "paper_no") > avg_no and _i(it, "paper_no") > 0:
            reasons.append(f"لا يوجد تقرير: {_i(it, 'paper_no')}")
        if _i(it, "paper_pending") > avg_pending and _i(it, "paper_pending") > 0:
            reasons.append(f"لم تجهز بعد: {_i(it, 'paper_pending')}")
        return reasons

    is_multi = n_tr > 1

    # ══════════════════════════════════════════════════════════════════════
    # صفحة 1 — الملخص التنفيذي (Executive Summary)
    # ══════════════════════════════════════════════════════════════════════
    if is_multi:
        story.append(HeaderBand(r("الملخص التنفيذي — تقييم أداء المترجمين"),
                                r(f"من {start_date_str} إلى {end_date_str}")))
        story.append(Spacer(1, 8))

        kpis = [
            _kpi("عدد المترجمين", str(n_tr), color=MAIN),
            _kpi("إجمالي التقارير", str(total_reports), color=MAIN),
            _kpi("إجمالي أيام الفترة", str(period_days_g), color=ACCENT),
            _kpi("إجمالي أيام العمل", str(tot_work_days), color=PURPLE),
            _kpi("قبل 8 مساء", str(tot_before8), color=ACCENT, value_color=GREEN),
            _kpi("بعد 8 مساء", str(total_late), color=RED, value_color=RED),
            _kpi("نسبة التقارير بعد 8", f"{late_pct_g:.0f}%", color=RED, value_color=RED),
            _kpi("تقارير تم رفعها", str(tot_yes), color=GREEN, value_color=GREEN),
            _kpi("نسبة رفع التقارير", f"{upload_pct_g:.0f}%", color=GREEN, value_color=GREEN),
            _kpi("تقارير لم تجهز بعد", str(tot_pending), color=AMBER, value_color=AMBER),
            _kpi("حالات لا يوجد تقرير", str(tot_no), color=RED, value_color=RED),
            _kpi("متوسط تقارير/مترجم", f"{(total_reports / n_tr):.1f}" if n_tr else "0", color=MAIN),
        ]
        story.append(_kpi_grid(kpis, per_row=3, cw=178))
        story.append(Spacer(1, 9))

        # رسمان دائريان جنباً إلى جنب
        pie_time = _pie([tot_before8, total_late], ["قبل 8 مساء", "بعد 8 مساء"],
                        [GREEN, RED], title="توزيع التقارير حسب التوقيت")
        pie_docs = _pie([tot_yes, tot_pending, tot_no],
                        ["تم رفعها", "لم تجهز بعد", "لا يوجد تقرير"],
                        [GREEN, AMBER, RED], title="حالة التقارير الطبية")
        if pie_time or pie_docs:
            row = [pie_docs or Drawing(262, 172), pie_time or Drawing(262, 172)]
            pt = Table([row], colWidths=[268, 268])
            pt.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (0, 0), 0.5, GRID),
                ("BOX", (1, 0), (1, 0), 0.5, GRID),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(pt)

        # ✅ رسم توزيع التقارير بين المترجمين هنا لملء الصفحة الأولى (بدل فراغ غير مبرر)
        top_n = ranked[:10]
        bar_tr = _hbar([str(it.get("translator_name", "—")) for it in top_n],
                       [_i(it, "total_reports") for it in top_n],
                       color=MAIN, width=536,
                       title="توزيع التقارير بين المترجمين (أعلى 10)")
        if bar_tr:
            story.append(Spacer(1, 9))
            story.append(bar_tr)
        story.append(PageBreak())

        # ══════════════════════════════════════════════════════════════════
        # صفحة 2 — جدول ترتيب المترجمين (نظرة إدارية شاملة)
        # ══════════════════════════════════════════════════════════════════
        story.append(HeaderBand(r("جدول ترتيب المترجمين"), r(f"إجمالي {n_tr} مترجم — {period_label}")))
        story.append(Spacer(1, 8))

        # RTL: أول عنصر منطقي يجب أن يكون آخر عمود (يظهر يميناً)
        rank_rows = [[r("لم تجهز"), r("لا يوجد"), r("مرفوعة"), r("نسبة التأخير"),
                      r("بعد 8"), r("قبل 8"), r("الإجمالي"), r("المترجم"), r("#")]]
        for idx, it in enumerate(ranked, 1):
            rank_rows.append([
                str(_i(it, "paper_pending")),
                str(_i(it, "paper_no")),
                str(_i(it, "paper_yes")),
                f"{_late_pct(it):.0f}%",
                str(_i(it, "late_reports")),
                str(_before8(it)),
                str(_i(it, "total_reports")),
                r(str(it.get("translator_name", "—"))),
                str(idx),
            ])
        rank_tbl = Table(rank_rows, colWidths=[54, 54, 54, 62, 48, 48, 54, 145, 40], repeatRows=1)
        rank_style = [
            ("BACKGROUND", (0, 0), (-1, 0), MAIN),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 9.5),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        # ✅ العمود رقم 3 هو «نسبة التأخير» في ترتيب الأعمدة أعلاه (وليس 5)
        _COL_LATE_PCT = 3
        _COL_RANK = 8
        for rr in range(1, len(rank_rows)):
            if rr % 2 == 0:
                rank_style.append(("BACKGROUND", (0, rr), (-1, rr), colors.HexColor("#FAFBFD")))
            it = ranked[rr - 1]
            if _i(it, "total_reports") > 0 and _late_pct(it) >= LATE_THRESHOLD:
                rank_style.append(("TEXTCOLOR", (_COL_LATE_PCT, rr), (_COL_LATE_PCT, rr), RED))
            if rr <= 3:
                rank_style.append(("TEXTCOLOR", (_COL_RANK, rr), (_COL_RANK, rr), MAIN))
        rank_tbl.setStyle(TableStyle(rank_style))
        story.append(rank_tbl)
        story.append(Spacer(1, 7))
        story.append(Paragraph(
            r(f"معيار «يحتاج متابعة»: نسبة تأخير ≥ {LATE_THRESHOLD:.0f}% أو «لا يوجد تقرير» / «لم تجهز بعد» أعلى من المتوسط."),
            ParagraphStyle("note", parent=styles["Normal"], fontName=font_name, fontSize=8.5,
                           leading=12, alignment=TA_CENTER, textColor=MUTED)))

        # ══════════════════════════════════════════════════════════════════
        # أبرز المؤشرات — تتدفق طبيعياً بعد الجدول (بلا فراغات غير مبررة)
        # ══════════════════════════════════════════════════════════════════
        story.append(Spacer(1, 13))
        story.append(Paragraph(r("أبرز المؤشرات"), section_style))
        story.append(Spacer(1, 6))

        def _hl(icon_label, it, value_text, color):
            name = str(it.get("translator_name", "—")) if it else "—"
            return _kpi(f"{icon_label} — {name}", value_text, color=color, value_color=color)

        # ملاحظة: قيمة البطاقة تُرسم بلا إعادة تشكيل عربي (كما في _draw_card الأصلية)،
        # لذا نُبقيها أرقاماً/نِسَباً فقط ونضع أي نص عربي في التسمية.
        hl_cells = [
            _hl("الأعلى عدد تقارير", best_volume, str(_i(best_volume, "total_reports")) if best_volume else "—", GREEN),
            _hl("الأكثر التزاماً (نسبة تأخير)", best_ontime, f"{_late_pct(best_ontime):.0f}%" if best_ontime else "—", GREEN),
            _hl("الأكثر رفعاً للتقارير", best_upload, str(_i(best_upload, "paper_yes")) if best_upload else "—", GREEN),
            _hl("أعلى نسبة تأخير", worst_late, f"{_late_pct(worst_late):.0f}%" if worst_late else "—", RED),
            _hl("أعلى «لا يوجد تقرير»", worst_no, str(_i(worst_no, "paper_no")) if worst_no else "—", RED),
            _hl("أعلى «لم تجهز بعد»", worst_pending, str(_i(worst_pending, "paper_pending")) if worst_pending else "—", AMBER),
        ]
        story.append(_kpi_grid(hl_cells, per_row=2, cw=266))
        story.append(Spacer(1, 12))

        agg_actions: dict[str, int] = {}
        agg_actors: dict[str, int] = {}   # عدد المترجمين الذين نفّذوا كل إجراء
        for it in results:
            for k, v in _norm_breakdown(it).items():
                agg_actions[k] = agg_actions.get(k, 0) + v
                agg_actors[k] = agg_actors.get(k, 0) + 1
        # ══════════════════════════════════════════════════════════════════
        # صفحة مستقلة: توزيع أنواع الإجراءات — رسم بياني + جدول إحصائيات
        # (الرسم وحده لا يعطي أرقاماً، فأُضيف الجدول بجانبه في نفس الصفحة)
        # ══════════════════════════════════════════════════════════════════
        all_actions = sorted(agg_actions.items(), key=lambda x: x[1], reverse=True)
        if all_actions:
            _sum_actions = sum(v for _, v in all_actions)
            story.append(PageBreak())
            story.append(HeaderBand(r("توزيع أنواع الإجراءات"),
                                    r(f"{period_label} — {_sum_actions} إجراء")))
            story.append(Spacer(1, 10))

            top_actions = all_actions[:8]
            bar_ac = _hbar([k for k, _ in top_actions], [v for _, v in top_actions],
                           color=ACCENT, width=534, title="الأعلى تكراراً")
            if bar_ac:
                story.append(bar_ac)
                story.append(Spacer(1, 10))

            # ✅ النسبة هنا = حصة الإجراء من إجمالي الإجراءات (تجمع دائماً 100%)
            # وهو المعنى الصحيح لجدول «توزيع». لم يُمَس اصطلاح جدول الإجراءات
            # الخاص بكل مترجم (الذي يقسم على إجمالي تقارير المترجم).
            act_rows = [[r("النسبة"), r("عدد المترجمين"), r("العدد"), r("نوع الإجراء")]]
            for _an, _ac in all_actions:
                act_rows.append([
                    f"{_pct(_ac, _sum_actions):.1f}%",
                    str(agg_actors.get(_an, 0)),
                    str(_ac),
                    r(_an),
                ])
            act_rows.append(["100%", "—", str(_sum_actions), r("الإجمالي")])

            act_tbl = Table(act_rows, colWidths=[84, 104, 74, 272], repeatRows=1)
            act_style = [
                ("BACKGROUND", (0, 0), (-1, 0), MAIN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 9.5),
                ("FONTSIZE", (0, 1), (-1, -1), 9.5),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("BACKGROUND", (0, -1), (-1, -1), MAIN_LIGHT),
                ("TEXTCOLOR", (0, -1), (-1, -1), MAIN),
            ]
            for _rr in range(1, len(act_rows) - 1):
                if _rr % 2 == 0:
                    act_style.append(("BACKGROUND", (0, _rr), (-1, _rr), colors.HexColor("#FAFBFD")))
                if all_actions[_rr - 1][0] == "تأجيل موعد":
                    act_style.append(("TEXTCOLOR", (0, _rr), (-1, _rr), RED))
            act_tbl.setStyle(TableStyle(act_style))
            story.append(act_tbl)
            story.append(Spacer(1, 7))
            story.append(Paragraph(
                r("النسبة = حصة الإجراء من إجمالي الإجراءات المنفَّذة خلال الفترة."),
                ParagraphStyle("actnote", parent=styles["Normal"], fontName=font_name,
                               fontSize=8.5, leading=12, alignment=TA_CENTER, textColor=MUTED)))

        story.append(PageBreak())

    for i, item in enumerate(results, 1):
        before_8pm = item.get("before_8pm_reports", max(item["total_reports"] - item["late_reports"], 0))
        total_reports_i = int(item.get("total_reports") or 0)
        late_reports_i = int(item.get("late_reports") or 0)
        late_pct = (late_reports_i / total_reports_i * 100.0) if total_reports_i > 0 else 0.0
        try:
            # _generate_pdf receives strings like "dd/mm/YYYY"
            _sd = datetime.strptime(start_date_str, "%d/%m/%Y").date()
            _ed = datetime.strptime(end_date_str, "%d/%m/%Y").date()
            period_days = max(int((_ed - _sd).days) + 1, 1)
        except Exception:
            period_days = "-"

        # ── صفحة 1: عنوان + فترة + إحصائيات ──
        story.append(HeaderBand(r(f"تقرير تقييم المترجم: {item['translator_name']}"), ""))
        story.append(Spacer(1, 4))
        story.append(Paragraph(r(f"من {start_date_str} إلى {end_date_str}"), date_style))
        story.append(Spacer(1, 6))

        # ✅ بطاقة أداء مختصرة قبل الكروت الحالية (الكروت الحالية تبقى كما هي تماماً)
        _ta_name, _ta_cnt = _top_action(item)
        _commit_pct = _pct(before_8pm, total_reports_i)
        _perf_rows = [[
            r("أكثر إجراء"), r("أيام العمل"), r("متوسط يومي"),
            r("نسبة التأخير"), r("نسبة الالتزام"), r("إجمالي التقارير"),
        ], [
            r(f"{_ta_name}" + (f" ({_ta_cnt})" if _ta_cnt else "")),
            str(item.get("work_days", 0)),
            f"{_avg_daily(item):.1f}",
            f"{late_pct:.0f}%",
            f"{_commit_pct:.0f}%",
            str(total_reports_i),
        ]]
        _perf = Table(_perf_rows, colWidths=[124, 74, 74, 84, 88, 86])
        _perf.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), MAIN_LIGHT),
            ("TEXTCOLOR", (0, 0), (-1, 0), MAIN),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, 1), 12),
            ("TEXTCOLOR", (3, 1), (3, 1), RED if late_pct >= LATE_THRESHOLD else INK),
            ("TEXTCOLOR", (4, 1), (4, 1), GREEN),
            ("TEXTCOLOR", (5, 1), (5, 1), MAIN),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(_perf)
        story.append(Spacer(1, 6))

        cards = Table(
            [
                # ترتيب RTL بصرياً: يمين أعلى (إجمالي) ثم يسار أعلى (قبل 8)
                [_draw_card("قبل 8 مساء", str(before_8pm), color=ACCENT), _draw_card("إجمالي التقارير", str(total_reports_i), color=MAIN)],
                # الصف الثاني: يمين (بعد 8) ثم يسار (نسبة المتأخر)
                [
                    _draw_card(
                        "بعد 8 مساء",
                        str(late_reports_i),
                        color=colors.HexColor("#EF5350"),
                        value_color=colors.HexColor("#D32F2F"),
                        label_color=colors.HexColor("#D32F2F"),
                    ),
                    _draw_card(
                        "نسبة التقارير بعد 8 مساء",
                        f"{late_pct:.0f}%",
                        color=colors.HexColor("#EF5350"),
                        value_color=colors.HexColor("#D32F2F"),
                        label_color=colors.HexColor("#D32F2F"),
                    ),
                ],
                # الصف الثالث: يمين (أيام العمل) ثم يسار (أيام الفترة) ليكونا البطاقة 5 و6
                [
                    _draw_card("إجمالي أيام الفترة", str(period_days), color=colors.HexColor("#26A69A")),
                    _draw_card("إجمالي أيام العمل", str(item["work_days"]), color=colors.HexColor("#7E57C2")),
                ],
                # الصف الرابع: لم تجهز بعد / تم رفعها
                # ملاحظة: أُزيلت رموز الإيموجي من التسميات فقط — الخط العربي
                # المستخدم في الـPDF لا يملك محارف إيموجي فكانت تُطبع مربعات
                # فارغة (□). التصميم والألوان والقيم كما هي بلا أي تغيير.
                [
                    _draw_card(
                        "لم تجهز بعد",
                        str(item.get("paper_pending", 0)),
                        color=colors.HexColor("#FFE0B2"),
                        value_color=colors.HexColor("#E65100"),
                        label_color=colors.HexColor("#E65100"),
                    ),
                    _draw_card(
                        "تم رفعها",
                        str(item.get("paper_yes", 0)),
                        color=colors.HexColor("#A5D6A7"),
                        value_color=colors.HexColor("#1B5E20"),
                        label_color=colors.HexColor("#1B5E20"),
                    ),
                ],
                # الصف الخامس: لا يوجد تقرير (البطاقة الثانية فارغة بلا إطار)
                [
                    _draw_card(
                        "لا يوجد تقرير",
                        str(item.get("paper_no", 0)),
                        color=colors.HexColor("#EF9A9A"),
                        value_color=colors.HexColor("#B71C1C"),
                        label_color=colors.HexColor("#B71C1C"),
                    ),
                    Drawing(250, 84),  # خانة فارغة (لا إطار) لموازنة الشبكة ثنائية الأعمدة
                ],
            ],
            colWidths=[260, 260],
        )
        cards.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(cards)

        # ── نفس الصفحة: جدول الإجراءات (دمج الصفحة الأولى والثانية) ──
        story.append(Spacer(1, 6))
        story.append(Paragraph(r("جدول الإجراءات"), section_style))
        story.append(Spacer(1, 4))

        raw_breakdown = item.get("action_breakdown", {})
        action_breakdown: dict[str, int] = {}
        for k, v in raw_breakdown.items():
            nk = _norm_action(k)
            if not nk:
                continue
            action_breakdown[nk] = action_breakdown.get(nk, 0) + int(v or 0)

        # الترتيب المطلوب: نوع الإجراء | العدد | النسبة
        # ملاحظة: ReportLab يرسم الأعمدة من اليسار لليمين، لذا نجعل "نوع الإجراء" عموداً أخيراً
        # ليظهر في أقصى اليمين أثناء القراءة العربية.
        rows = [[r("النسبة"), r("العدد"), r("نوع الإجراء")]]
        reschedule_row_idx = None
        for action_name, count in sorted(action_breakdown.items(), key=lambda x: x[1], reverse=True):
            if count <= 0:
                continue
            pct = (count / item["total_reports"] * 100.0) if item["total_reports"] > 0 else 0.0
            rows.append([f"{pct:.0f}%", str(count), r(action_name)])
            if action_name == "تأجيل موعد":
                reschedule_row_idx = len(rows) - 1

        if len(rows) == 1:
            story.append(Paragraph(r("لا توجد بيانات إجراءات لعرضها"), base_style))
        else:
            tbl = Table(rows, colWidths=[90, 70, 360], repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), MAIN),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 16),
                ("FONTSIZE", (0, 1), (-1, -1), 14),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID),
                # توسيط النصوص والأرقام داخل الجدول
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                # رفع المحتوى للأعلى داخل الخلايا
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, 0), 3),   # الهيدر أعلى
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 2),  # الصفوف أعلى
                ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
            ]
            for rr in range(1, len(rows)):
                if rr % 2 == 0:
                    style_cmds.append(("BACKGROUND", (0, rr), (-1, rr), colors.HexColor("#FAFBFD")))
            # تمييز "تأجيل موعد" باللون الأحمر
            if reschedule_row_idx is not None:
                style_cmds.append(("TEXTCOLOR", (0, reschedule_row_idx), (-1, reschedule_row_idx), colors.HexColor("#D32F2F")))
                style_cmds.append(("FONTSIZE", (0, reschedule_row_idx), (-1, reschedule_row_idx), 13))
            tbl.setStyle(TableStyle(style_cmds))
            story.append(tbl)

        story.append(PageBreak())

        # ── صفحة 3: النشاط اليومي (رسم بياني) ──
        story.append(Paragraph(r(f"{item['translator_name']} - النشاط اليومي"), section_style))
        story.append(Spacer(1, 4))
        item_start = item.get("start_date")
        item_end = item.get("end_date")
        daily = []
        if item_start and item_end:
            try:
                from db.session import SessionLocal
                with SessionLocal() as pdf_session:
                    daily = _get_daily_counts(pdf_session, item.get("translator_id"), item_start, item_end, translator_name=item.get("translator_name"))
            except Exception as e:
                logger.warning(f"خطأ في بيانات النشاط اليومي للـ PDF: {e}")

        if daily:
            labels = []
            values = []
            # تقليل كثافة العلامات للحفاظ على وضوح الرسم
            sampled = daily[-20:] if len(daily) > 20 else daily
            for d, c in sampled:
                try:
                    dt = datetime.strptime(d, "%Y-%m-%d")
                    labels.append(dt.strftime("%d/%m"))
                except Exception:
                    labels.append(d)
                values.append(int(c))

            # توسيط الرسم البياني عمودياً داخل الصفحة قدر الإمكان
            story.append(Spacer(1, 90))
            drawing = Drawing(520, 250)
            chart = VerticalBarChart()
            chart.x = 40
            chart.y = 38
            chart.height = 180
            chart.width = 450
            chart.data = [values]
            chart.strokeColor = GRID
            chart.valueAxis.valueMin = 0
            chart.valueAxis.valueMax = max(values) + 1
            chart.valueAxis.valueStep = max(1, int((max(values) + 1) / 5))
            chart.categoryAxis.categoryNames = labels
            chart.categoryAxis.labels.angle = 35
            chart.categoryAxis.labels.dy = -12
            chart.categoryAxis.labels.fontName = font_name
            chart.categoryAxis.labels.fontSize = 8
            chart.valueAxis.labels.fontName = font_name
            chart.valueAxis.labels.fontSize = 8
            chart.bars[0].fillColor = MAIN
            chart.bars[0].strokeColor = MAIN
            drawing.add(Rect(0, 0, 520, 250, fillColor=colors.HexColor("#FFFFFF"), strokeColor=GRID, strokeWidth=0.7))
            drawing.add(chart)
            story.append(drawing)
            story.append(Spacer(1, 12))

            total_day_tbl = Table(
                [[r("عدد الأيام المعروضة"), str(len(sampled)), r("أعلى نشاط يومي"), str(max(values))]],
                colWidths=[130, 100, 130, 100],
            )
            total_day_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), MAIN_LIGHT),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ]))
            story.append(total_day_tbl)
            story.append(Spacer(1, 6))

            # جدول كامل لكل الأيام: التاريخ | عدد التقارير
            # لتفادي صفحة أخيرة فيها سطر/سطرين فقط، نفصل الجدول بصفحة مستقلة إذا كانت الأيام كثيرة.
            show_table_on_new_page = len(daily) > 15
            if show_table_on_new_page:
                story.append(PageBreak())

            day_rows = [[r("عدد التقارير"), r("التاريخ")]]
            for day_str, count in daily:
                day_rows.append([str(count), r(day_str)])
            day_table = Table(day_rows, colWidths=[120, 380], repeatRows=1)
            day_style = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 15),
                ("FONTSIZE", (0, 1), (-1, -1), 14),
                ("GRID", (0, 0), (-1, -1), 0.35, GRID),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                # رفع التاريخ وعدد التقارير للأعلى داخل الخلايا
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, 0), 3),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 7),
                ("TOPPADDING", (0, 1), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 7),
            ]
            for rr in range(1, len(day_rows)):
                if rr % 2 == 0:
                    day_style.append(("BACKGROUND", (0, rr), (-1, rr), colors.HexColor("#FAFBFD")))
            day_table.setStyle(TableStyle(day_style))
            story.append(Paragraph(r("تفاصيل الأيام (التاريخ | عدد التقارير)"), base_style))
            story.append(Spacer(1, 3))
            story.append(day_table)
        else:
            story.append(Paragraph(r("لا توجد بيانات نشاط يومي متاحة لهذه الفترة."), base_style))

        # ══════════════════════════════════════════════════════════════════
        # صفحة إضافية لكل مترجم: رسوم تحليلية + ملخص نقاط القوة/التحسين
        # (إضافة فقط — لا تحذف أو تغيّر أي رسم أو جدول قائم)
        # ══════════════════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph(r(f"{item['translator_name']} — تحليل الأداء"), section_style))
        story.append(Spacer(1, 6))

        # النشاط الأسبوعي (مشتق من نفس بيانات daily المجلوبة أصلاً)
        if daily:
            _wd_names = ["الاثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
            _wk = [0] * 7
            for _d, _c in daily:
                try:
                    _wk[datetime.strptime(_d, "%Y-%m-%d").weekday()] += int(_c)
                except Exception:
                    pass
            if sum(_wk) > 0:
                wdraw = Drawing(536, 190)
                wdraw.add(String(530, 176, r("النشاط الأسبوعي (مجموع التقارير لكل يوم)"),
                                 fontName=font_name, fontSize=10.5, fillColor=INK, textAnchor="end"))
                wch = VerticalBarChart()
                wch.x = 36
                wch.y = 34
                wch.width = 470
                wch.height = 126
                wch.data = [_wk]
                wch.strokeColor = GRID
                wch.valueAxis.valueMin = 0
                wch.valueAxis.valueMax = max(_wk) + 1
                wch.valueAxis.valueStep = max(1, int((max(_wk) + 1) / 4))
                wch.valueAxis.labels.fontName = font_name
                wch.valueAxis.labels.fontSize = 8
                wch.categoryAxis.categoryNames = [r(x) for x in _wd_names]
                wch.categoryAxis.labels.fontName = font_name
                wch.categoryAxis.labels.fontSize = 8
                wch.bars[0].fillColor = PURPLE
                wch.bars[0].strokeColor = PURPLE
                wdraw.add(Rect(0, 0, 536, 190, fillColor=colors.white, strokeColor=GRID, strokeWidth=0.7))
                wdraw.add(wch)
                story.append(wdraw)
                story.append(Spacer(1, 8))

        # توزيع التوقيت (دائري) + توزيع الإجراءات (أعمدة أفقية) جنباً إلى جنب
        _pie_i = _pie([before_8pm, late_reports_i], ["قبل 8 مساء", "بعد 8 مساء"],
                      [GREEN, RED], width=250, height=164, title="توزيع التوقيت")
        _bd_i = _norm_breakdown(item)
        _top_i = sorted(_bd_i.items(), key=lambda x: x[1], reverse=True)[:6]
        _bar_i = _hbar([k for k, _ in _top_i], [v for _, v in _top_i],
                       color=ACCENT, width=274, height=164, title="أنواع الإجراءات") if _top_i else None
        if _pie_i or _bar_i:
            _row = [_bar_i or Drawing(274, 164), _pie_i or Drawing(250, 164)]
            _t = Table([_row], colWidths=[276, 258])
            _t.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (0, 0), 0.5, GRID),
                ("BOX", (1, 0), (1, 0), 0.5, GRID),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(_t)
            story.append(Spacer(1, 9))

        # ── ملخص الأداء: نقاط القوة / نقاط تحتاج تحسين (من البيانات فقط) ──
        _pyes, _pno, _ppend = _i(item, "paper_yes"), _i(item, "paper_no"), _i(item, "paper_pending")
        _wd = _i(item, "work_days")
        _adaily = _avg_daily(item)
        _pd = period_days_g if isinstance(period_days_g, int) else 0

        strengths = []
        if total_reports_i > 0 and late_pct <= 20:
            strengths.append(f"التزام ممتاز بالوقت (نسبة تأخير {late_pct:.0f}%)")
        if total_reports_i > 0 and _commit_pct >= 80:
            strengths.append(f"معظم التقارير قبل 8 مساء ({_commit_pct:.0f}%)")
        if _pyes > 0:
            strengths.append(f"رفع {_pyes} تقرير طبي")
        if team_avg_daily > 0 and _adaily >= team_avg_daily:
            strengths.append(f"معدل يومي ({_adaily:.1f}) أعلى من متوسط الفريق ({team_avg_daily:.1f})")
        if _pd and _wd >= 0.7 * _pd:
            strengths.append(f"انتظام حضور مرتفع ({_wd} من {_pd} يوم)")
        if _ta_cnt:
            strengths.append(f"أكثر إجراء: {_ta_name} ({_ta_cnt})")

        needs = []
        if total_reports_i > 0 and late_pct >= LATE_THRESHOLD:
            needs.append(f"نسبة تأخير مرتفعة ({late_pct:.0f}%)")
        if _pno > 0:
            needs.append(f"{_pno} حالة بلا تقرير طبي")
        if _ppend > 0:
            needs.append(f"{_ppend} تقرير لم يجهز بعد")
        if team_avg_daily > 0 and _adaily < team_avg_daily:
            needs.append(f"معدل يومي ({_adaily:.1f}) أقل من متوسط الفريق ({team_avg_daily:.1f})")
        if _pd and _wd < 0.4 * _pd:
            needs.append(f"أيام عمل منخفضة ({_wd} من {_pd} يوم)")
        if not needs:
            needs.append("لا توجد ملاحظات — الأداء ضمن المعدلات")

        _sw = Table([[_bullets("نقاط تحتاج تحسين", needs[:5], RED),
                      _bullets("نقاط القوة", strengths[:5], GREEN)]],
                    colWidths=[268, 268])
        _sw.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(_sw)

        if i < len(results):
            story.append(PageBreak())

    # ══════════════════════════════════════════════════════════════════════
    # الصفحة الختامية — الخلاصة العامة + يحتاج متابعة
    # ══════════════════════════════════════════════════════════════════════
    if is_multi:
        story.append(PageBreak())
        story.append(HeaderBand(r("الخلاصة العامة"), r(f"{period_label} — {n_tr} مترجم — {total_reports} تقرير")))
        story.append(Spacer(1, 10))

        def _nm(it):
            return str(it.get("translator_name", "—")) if it else "—"

        # القيم أرقام/نِسَب فقط (بلا نص عربي) — أي وصف عربي يوضَع في عمود «المؤشر»
        summary_rows = [[r("القيمة"), r("المترجم"), r("المؤشر")]]
        summary_rows += [
            [str(_i(best_volume, "total_reports")) if best_volume else "—", r(_nm(best_volume)), r("أفضل مترجم (الأعلى عدد تقارير)")],
            [f"{_late_pct(best_ontime):.0f}%" if best_ontime else "—", r(_nm(best_ontime)), r("الأكثر التزاماً بالوقت (نسبة تأخير)")],
            [f"{_avg_daily(best_volume):.1f}" if best_volume else "—", r(_nm(best_volume)), r("الأعلى نشاطاً (متوسط تقارير/يوم)")],
            [str(_i(best_upload, "paper_yes")) if best_upload else "—", r(_nm(best_upload)), r("الأكثر رفعاً للتقارير الطبية")],
            [f"{_late_pct(best_ontime):.0f}%" if best_ontime else "—", r(_nm(best_ontime)), r("الأقل تأخيراً (نسبة تأخير)")],
        ]
        s_tbl = Table(summary_rows, colWidths=[150, 176, 210], repeatRows=1)
        s_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), MAIN),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTSIZE", (0, 1), (-1, -1), 10.5),
            ("TEXTCOLOR", (0, 1), (0, -1), GREEN),
            ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ]))
        for _rr in range(1, len(summary_rows)):
            if _rr % 2 == 0:
                s_tbl.setStyle(TableStyle([("BACKGROUND", (0, _rr), (-1, _rr), colors.HexColor("#FAFBFD"))]))
        story.append(s_tbl)
        story.append(Spacer(1, 16))

        story.append(Paragraph(r("يحتاج متابعة"), section_style))
        story.append(Spacer(1, 5))
        follow = [(it, _needs_followup(it)) for it in ranked]
        follow = [(it, rs) for it, rs in follow if rs]
        if not follow:
            story.append(Paragraph(r("لا يوجد مترجمون يحتاجون متابعة وفق المعايير المحددة. ✔"), base_style))
        else:
            f_rows = [[r("سبب المتابعة"), r("المترجم")]]
            for it, rs in follow:
                f_rows.append([r("، ".join(rs)), r(_nm(it))])
            f_tbl = Table(f_rows, colWidths=[360, 176], repeatRows=1)
            f_style = [
                ("BACKGROUND", (0, 0), (-1, 0), RED),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("FONTSIZE", (0, 1), (-1, -1), 9.5),
                ("GRID", (0, 0), (-1, -1), 0.4, GRID),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
            for _rr in range(1, len(f_rows)):
                if _rr % 2 == 0:
                    f_style.append(("BACKGROUND", (0, _rr), (-1, _rr), colors.HexColor("#FFF5F5")))
            f_tbl.setStyle(TableStyle(f_style))
            story.append(f_tbl)

        story.append(Spacer(1, 14))
        story.append(Paragraph(
            r(f"معايير المتابعة: نسبة تأخير ≥ {LATE_THRESHOLD:.0f}% • «لا يوجد تقرير» أعلى من المتوسط ({avg_no:.1f}) • «لم تجهز بعد» أعلى من المتوسط ({avg_pending:.1f})"),
            ParagraphStyle("crit", parent=styles["Normal"], fontName=font_name, fontSize=8.5,
                           leading=12, alignment=TA_CENTER, textColor=MUTED)))

    doc.build(story)
    return buffer.getvalue(), "pdf"


def _generate_html_fallback(results, period_label, year, month, start_date_str, end_date_str, total_reports, total_late, target_name: str | None = None):
    """HTML fallback إذا فشل reportlab"""
    total_before = total_reports - total_late
    translator_pages = ""
    for i, item in enumerate(results, 1):
        actions_rows = ""
        for action_name, count in sorted(item.get('action_breakdown', {}).items(), key=lambda x: x[1], reverse=True):
            color = "" if count > 0 else ' style="color:#bbb;"'
            actions_rows += f'<tr{color}><td style="text-align:right;padding:7px 12px;font-size:14px;">{action_name}</td><td style="text-align:center;padding:7px 12px;font-size:14px;">{count}</td></tr>'
        before_8pm = item.get("before_8pm_reports", max(item["total_reports"] - item["late_reports"], 0))
        paper_yes_i = item.get("paper_yes", 0)
        paper_no_i = item.get("paper_no", 0)
        paper_pending_i = item.get("paper_pending", 0)
        rank_prefix = "" if target_name else f"{_medal(i)} "
        translator_pages += f'''<div style="page-break-before:always;font-size:15px;"><h2 style="font-size:22px;">{rank_prefix}{item["translator_name"]}</h2>
        <p style="font-size:16px;">إجمالي التقارير: <b>{item["total_reports"]}</b> | قبل 8 مساءً: <b>{before_8pm}</b> | بعد 8 مساءً: <b>{item["late_reports"]}</b> | أيام العمل: <b>{item["work_days"]}</b> | ✅ تم رفعها: <b>{paper_yes_i}</b> | 🟡 لم تجهز بعد: <b>{paper_pending_i}</b> | ❌ لا يوجد تقرير: <b>{paper_no_i}</b></p>
        <table border="1" cellpadding="5" style="font-size:14px;border-collapse:collapse;"><tr style="background:#1a237e;color:#fff;font-size:15px;"><th style="padding:8px 12px;">نوع الإجراء</th><th style="padding:8px 12px;">العدد</th></tr>{actions_rows}</table></div>'''

    header_title = "تقرير تقييم مترجم فردي" if target_name else "تقرير تقييم المترجمين"
    sub = f"المترجم: <b>{target_name}</b> | " if target_name else ""
    summary = (
        f"{sub}{period_label} | تقارير: <b>{total_reports}</b> | قبل 8 مساءً: <b>{total_before}</b> | بعد 8 مساءً: <b>{total_late}</b>"
        if target_name else
        f"{period_label} | مترجمين: <b>{len(results)}</b> | تقارير: <b>{total_reports}</b> | قبل 8 مساءً: <b>{total_before}</b> | بعد 8 مساءً: <b>{total_late}</b>"
    )
    html = (
        f'<!DOCTYPE html><html dir="rtl" lang="ar"><head><meta charset="UTF-8">'
        f'<style>body{{font-family:Tahoma,Arial,sans-serif;font-size:15px;}}h1{{font-size:26px;}}p{{font-size:16px;}}</style>'
        f'</head><body><h1>{header_title}</h1><p>{summary}</p>{translator_pages}</body></html>'
    )
    return html.encode("utf-8")


# ════════════════════════════════════════
# توليد ملف Excel
# ════════════════════════════════════════

def _generate_excel(results, period_label, year, month, target_name: str | None = None):
    """توليد تقرير Excel - ملخص + تفصيل الإجراءات"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties

    wb = Workbook()

    # ─── الورقة 1: ملخص ───
    ws = wb.active
    ws.title = "ملخص التقييمات"
    ws.sheet_view.rightToLeft = True
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    title_font = Font(name='Arial', bold=True, size=14, color='1A237E')
    bold_font = Font(name='Arial', bold=True, size=11)
    normal_font = Font(name='Arial', size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    ws.merge_cells('A1:I1')
    main_title = (
        f"تقرير تقييم مترجم فردي - {target_name} - {period_label}"
        if target_name else f"تقرير تقييم أداء المترجمين - {period_label}"
    )
    ws['A1'] = main_title
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:I2')
    ws['A2'] = f"تاريخ الإصدار: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, color='777777')
    ws['A2'].alignment = center_align

    headers = ['الترتيب', 'المترجم', 'إجمالي التقارير', 'قبل 8 مساءً', 'بعد 8 مساءً', 'أيام العمل', '✅ تم رفعها', '🟡 لم تجهز بعد', '❌ لا يوجد تقرير']

    row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, item in enumerate(results, 1):
        row = i + 4
        medal = ""
        if i == 1: medal = "🥇 "
        elif i == 2: medal = "🥈 "
        elif i == 3: medal = "🥉 "

        before_8pm = item.get('before_8pm_reports', max(item['total_reports'] - item['late_reports'], 0))
        values = [
            i,
            f"{medal}{item['translator_name']}",
            item['total_reports'],
            before_8pm,
            item['late_reports'],
            item['work_days'],
            item.get('paper_yes', 0),
            item.get('paper_pending', 0),
            item.get('paper_no', 0),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = bold_font if col == 2 else normal_font
            cell.alignment = center_align if col != 2 else right_align
            cell.border = thin_border

    col_widths = [8, 28, 15, 14, 14, 12, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── الورقة 2: تفصيل الإجراءات ───
    ws2 = wb.create_sheet("تفصيل الإجراءات")
    ws2.sheet_view.rightToLeft = True
    ws2.page_setup.orientation = 'landscape'
    ws2.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws2.print_options.horizontalCentered = True

    total_cols = 1 + len(ALL_ACTION_TYPES) + 3
    end_col_letter = get_column_letter(total_cols)
    ws2.merge_cells(f'A1:{end_col_letter}1')
    ws2['A1'] = f"تفصيل التقارير حسب نوع الإجراء - {period_label}"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center_align

    detail_headers = ['المترجم'] + ALL_ACTION_TYPES + ['المجموع', 'قبل 8 مساءً', 'بعد 8 مساءً']

    row = 3
    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, item in enumerate(results, 1):
        row = i + 3
        action_breakdown = item.get('action_breakdown', {})
        values = [item['translator_name']]
        for action_type in ALL_ACTION_TYPES:
            values.append(action_breakdown.get(action_type, 0))
        before_8pm = item.get('before_8pm_reports', max(item['total_reports'] - item['late_reports'], 0))
        values.append(item['total_reports'])
        values.append(before_8pm)
        values.append(item['late_reports'])

        for col, val in enumerate(values, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.font = normal_font
            cell.alignment = center_align if col != 1 else right_align
            cell.border = thin_border

    ws2.column_dimensions['A'].width = 25
    for col_idx in range(2, total_cols + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ════════════════════════════════════════
# Handlers
# ════════════════════════════════════════

def _build_mode_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 تقييم جميع المترجمين", callback_data="evalmode:all")],
        [InlineKeyboardButton("👤 تقييم مترجم فردي", callback_data="evalmode:single")],
        [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
    ])


def _fetch_translators_for_picker() -> list[dict]:
    with SessionLocal() as s:
        rows = s.query(TranslatorDirectory).order_by(TranslatorDirectory.name).all()
        return [
            {"id": r.translator_id, "name": (r.name or "").strip()}
            for r in rows
            if r.name and r.name.strip()
        ]


def _build_translators_picker(page: int) -> InlineKeyboardMarkup:
    translators = _fetch_translators_for_picker()
    total = len(translators)
    per_page = EVAL_TRANSLATORS_PER_PAGE
    pages = max(1, (total + per_page - 1) // per_page)
    page = max(0, min(page, pages - 1))
    start = page * per_page
    end = min(start + per_page, total)

    keyboard: list[list[InlineKeyboardButton]] = []
    for t in translators[start:end]:
        keyboard.append([InlineKeyboardButton(
            f"👤 {t['name']}",
            callback_data=f"evaltr:pick:{t['id'] or 0}"
        )])

    nav: list[InlineKeyboardButton] = []
    if pages > 1:
        if page > 0:
            nav.append(InlineKeyboardButton("◀️ السابق", callback_data=f"evaltr:page:{page - 1}"))
        nav.append(InlineKeyboardButton(f"{page + 1}/{pages}", callback_data="noop"))
        if page < pages - 1:
            nav.append(InlineKeyboardButton("التالي ▶️", callback_data=f"evaltr:page:{page + 1}"))
    if nav:
        keyboard.append(nav)

    keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="evalmode:back")])
    keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])
    return InlineKeyboardMarkup(keyboard)


def _build_years_keyboard() -> InlineKeyboardMarkup:
    current_year = date.today().year
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📅 {current_year}", callback_data=f"eval:year:{current_year}")],
        [InlineKeyboardButton(f"📅 {current_year - 1}", callback_data=f"eval:year:{current_year - 1}")],
        [InlineKeyboardButton("🔙 رجوع", callback_data="evalmode:back")],
        [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
    ])


def _eval_intro_text() -> str:
    return (
        "╔══════════════════════════════════╗\n"
        "     📊 **تقييم أداء المترجمين**\n"
        "╚══════════════════════════════════╝\n\n"
        "📌 **التقرير يتضمن:**\n"
        "├ 👤 اسم المترجم\n"
        "├ 📅 الفترة (من - إلى)\n"
        "├ 📄 إجمالي التقارير\n"
        "├ 🌇 قبل 8 مساءً\n"
        "├ 🌙 بعد 8 مساءً\n"
        "├ 📅 عدد أيام العمل\n"
        "├ ✅ تقارير تم رفعها\n"
        "├ 🟡 تقارير لم تجهز بعد\n"
        "├ ❌ حالات لا يوجد لها تقرير\n"
        "└ 📋 تفصيل حسب نوع الإجراء\n\n"
        "اختر نوع التقييم:"
    )


async def handle_eval_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    data = q.data
    if data == "evalmode:back":
        await q.edit_message_text(
            _eval_intro_text(),
            reply_markup=_build_mode_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_MODE

    eval_data = context.user_data.setdefault('eval_data', {})

    if data == "evalmode:all":
        eval_data['mode'] = 'all'
        eval_data.pop('target_translator_name', None)
        eval_data.pop('target_translator_id', None)
        await q.edit_message_text(
            "📊 **تقييم أداء المترجمين**\n\n👥 الوضع: **كل المترجمين**\n\nاختر السنة:",
            reply_markup=_build_years_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_YEAR

    if data == "evalmode:single":
        eval_data['mode'] = 'single'
        await q.edit_message_text(
            "👤 **تقييم مترجم فردي**\n\nاختر المترجم من القائمة:",
            reply_markup=_build_translators_picker(0),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_TRANSLATOR

    return EVAL_SELECT_MODE


async def handle_translator_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    data = q.data
    if data == "evalmode:back":
        await q.edit_message_text(
            _eval_intro_text(),
            reply_markup=_build_mode_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_MODE

    if data.startswith("evaltr:page:"):
        page = int(data.split(":")[2])
        try:
            await q.edit_message_reply_markup(reply_markup=_build_translators_picker(page))
        except Exception:
            await q.edit_message_text(
                "👤 **تقييم مترجم فردي**\n\nاختر المترجم من القائمة:",
                reply_markup=_build_translators_picker(page),
                parse_mode=ParseMode.MARKDOWN,
            )
        return EVAL_SELECT_TRANSLATOR

    if data.startswith("evaltr:pick:"):
        try:
            tid = int(data.split(":")[2])
        except ValueError:
            tid = 0
        with SessionLocal() as s:
            tr = s.query(TranslatorDirectory).filter_by(translator_id=tid).first() if tid else None
            name = (tr.name or "").strip() if tr else ""
        if not name:
            await q.answer("⚠️ تعذّر قراءة المترجم", show_alert=True)
            return EVAL_SELECT_TRANSLATOR

        eval_data = context.user_data.setdefault('eval_data', {})
        eval_data['target_translator_id'] = tid or None
        eval_data['target_translator_name'] = name

        await q.edit_message_text(
            f"👤 **تقييم مترجم فردي**\n\nالمترجم: **{name}**\n\nاختر السنة:",
            reply_markup=_build_years_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_YEAR

    return EVAL_SELECT_TRANSLATOR


async def start_evaluation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """نقطة الدخول - اختيار وضع التقييم"""
    user = update.effective_user
    if not is_admin(user.id):
        await update.message.reply_text("هذه الخاصية مخصصة للأدمن فقط.")
        return ConversationHandler.END

    context.user_data.pop('eval_data', None)

    await update.message.reply_text(
        _eval_intro_text(),
        reply_markup=_build_mode_keyboard(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EVAL_SELECT_MODE


def _build_day_calendar(year, month):
    today = date.today()
    keyboard = []
    keyboard.append([InlineKeyboardButton(f"{MONTHS_AR[month]} {year}", callback_data="noop")])
    keyboard.append([InlineKeyboardButton(day, callback_data="noop") for day in DAYS_AR])

    for week in calendar.monthcalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="noop"))
                continue
            day_date = date(year, month, day)
            if day_date > today:
                row.append(InlineKeyboardButton(f"·{day}·", callback_data="noop"))
            else:
                date_str = f"{year}-{month:02d}-{day:02d}"
                row.append(InlineKeyboardButton(str(day), callback_data=f"evalday:select:{date_str}"))
        keyboard.append(row)

    keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_period")])
    keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])
    return InlineKeyboardMarkup(keyboard)


def _build_custom_calendar(year, month, step="start"):
    """بناء تقويم مخصص لاختيار البداية/النهاية"""
    today = date.today()
    keyboard = []
    
    # عنوان الشهر والسنة مع أزرار التنقل
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    keyboard.append([
        InlineKeyboardButton("◀️", callback_data=f"evalcustom:nav:{prev_year}:{prev_month}:{step}"),
        InlineKeyboardButton(f"{MONTH_NAMES[month]} {year}", callback_data="noop"),
        InlineKeyboardButton("▶️", callback_data=f"evalcustom:nav:{next_year}:{next_month}:{step}")
    ])
    
    keyboard.append([InlineKeyboardButton(day, callback_data="noop") for day in DAYS_AR])

    for week in calendar.monthcalendar(year, month):
        row = []
        for day in week:
            if day == 0:
                row.append(InlineKeyboardButton(" ", callback_data="noop"))
                continue
            day_date = date(year, month, day)
            if day_date > today:
                row.append(InlineKeyboardButton(f"·{day}·", callback_data="noop"))
            else:
                date_str = f"{year}-{month:02d}-{day:02d}"
                row.append(InlineKeyboardButton(str(day), callback_data=f"evalcustom:select:{date_str}:{step}"))
        keyboard.append(row)

    if step == "start":
        keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_year")])
    else:
        keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:month:custom")])
    keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])
    return InlineKeyboardMarkup(keyboard)


async def handle_custom_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """معالجة اختيار التاريخ في الفترة المخصصة"""
    q = update.callback_query
    await q.answer()
    
    data = q.data.split(":")
    action = data[1]
    
    eval_data = context.user_data.setdefault('eval_data', {})
    
    if action == "nav":
        y, m, step = int(data[2]), int(data[3]), data[4]
        title = "تاريخ البداية" if step == "start" else "تاريخ النهاية"
        await q.edit_message_text(
            f"📅 **تقييم فترة مخصصة**\n\nاختر **{title}**:",
            reply_markup=_build_custom_calendar(y, m, step),
            parse_mode=ParseMode.MARKDOWN
        )
        return EVAL_CUSTOM_START if step == "start" else EVAL_CUSTOM_END
        
    if action == "select":
        date_str, step = data[2], data[3]
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        
        if step == "start":
            eval_data['start_date'] = selected_date
            # الانتقال لاختيار تاريخ النهاية
            await q.edit_message_text(
                f"📅 **تقييم فترة مخصصة**\n\n"
                f"تاريخ البداية: **{selected_date}**\n\n"
                f"الآن اختر **تاريخ النهاية**:",
                reply_markup=_build_custom_calendar(selected_date.year, selected_date.month, "end"),
                parse_mode=ParseMode.MARKDOWN
            )
            return EVAL_CUSTOM_END
        else:
            start_date = eval_data.get('start_date')
            if start_date and selected_date < start_date:
                await q.answer("⚠️ تاريخ النهاية لا يمكن أن يكون قبل تاريخ البداية", show_alert=True)
                return EVAL_CUSTOM_END
                
            eval_data['end_date'] = selected_date
            period_label = f"من {start_date} إلى {selected_date}"
            
            keyboard = [
                [InlineKeyboardButton("📄 PDF", callback_data="eval:format:pdf")],
                [InlineKeyboardButton("📊 Excel", callback_data="eval:format:excel")],
                [InlineKeyboardButton("📄 PDF + 📊 Excel", callback_data="eval:format:both")],
                [InlineKeyboardButton("🔙 رجوع", callback_data="eval:month:custom")],
                [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
            ]
            
            await q.edit_message_text(
                f"📊 **تقييم أداء المترجمين**\n\n"
                f"📅 الفترة: **{period_label}**\n\n"
                f"اختر صيغة الملف:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode=ParseMode.MARKDOWN,
            )
            return EVAL_SELECT_FORMAT

    return EVAL_CUSTOM_START


async def handle_year(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """اختيار السنة"""
    q = update.callback_query
    await q.answer()

    if q.data == "eval:cancel":
        await q.edit_message_text("✅ تم إلغاء التقييم.")
        return ConversationHandler.END

    year = int(q.data.split(":")[2])
    context.user_data.setdefault('eval_data', {})['year'] = year

    keyboard = []
    for i in range(0, 12, 3):
        row = []
        for j in range(3):
            m = i + j + 1
            row.append(InlineKeyboardButton(
                MONTH_NAMES[m], callback_data=f"eval:month:{m}"
            ))
        keyboard.append(row)

    keyboard.append([InlineKeyboardButton("📄 كل الشهور", callback_data="eval:month:all")])
    keyboard.append([InlineKeyboardButton("📅 فترة مخصصة (من - إلى)", callback_data="eval:month:custom")])
    keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_year")])
    keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])

    await q.edit_message_text(
        f"📊 **تقييم أداء المترجمين**\n\n"
        f"📅 السنة: **{year}**\n\n"
        f"اختر الشهر:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EVAL_SELECT_MONTH


async def handle_month(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """اختيار الشهر"""
    q = update.callback_query
    await q.answer()

    if q.data == "eval:cancel":
        await q.edit_message_text("✅ تم إلغاء التقييم.")
        return ConversationHandler.END

    if q.data == "eval:back_year":
        eval_data = context.user_data.get('eval_data', {})
        mode = eval_data.get('mode', 'all')
        tname = (eval_data.get('target_translator_name') or "").strip()
        if mode == 'single' and tname:
            title = f"👤 **تقييم مترجم فردي**\n\nالمترجم: **{tname}**\n\nاختر السنة:"
        else:
            title = "📊 **تقييم أداء المترجمين**\n\n👥 الوضع: **كل المترجمين**\n\nاختر السنة:"
        await q.edit_message_text(
            title,
            reply_markup=_build_years_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_YEAR

    month_val = q.data.split(":")[2]
    eval_data = context.user_data.setdefault('eval_data', {})
    eval_data['month'] = month_val
    eval_data.pop('day', None)
    eval_data.pop('period_type', None)
    eval_data.pop('start_date', None)
    eval_data.pop('end_date', None)

    year = eval_data.get('year', date.today().year)
    
    if month_val == "custom":
        eval_data['period_type'] = "custom"
        # عرض تقويم لاختيار تاريخ البداية
        await q.edit_message_text(
            "📅 **تقييم فترة مخصصة**\n\n"
            "الخطوة 1: اختر **تاريخ البداية**:",
            reply_markup=_build_custom_calendar(year, date.today().month, "start"),
            parse_mode=ParseMode.MARKDOWN
        )
        return EVAL_CUSTOM_START

    month_label = "كل الشهور" if month_val == "all" else MONTH_NAMES.get(int(month_val), month_val)

    if month_val == "all":
        eval_data['period_type'] = "full"
        keyboard = [
            [InlineKeyboardButton("📄 PDF", callback_data="eval:format:pdf")],
            [InlineKeyboardButton("📊 Excel", callback_data="eval:format:excel")],
            [InlineKeyboardButton("📄 PDF + 📊 Excel", callback_data="eval:format:both")],
            [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
        ]

        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n"
            f"📅 الفترة: **{month_label} {year}**\n\n"
            f"اختر صيغة الملف:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_FORMAT

    keyboard = [
        [InlineKeyboardButton("📅 الشهر كامل", callback_data="eval:period:full")],
        [InlineKeyboardButton("📆 يوم محدد", callback_data="eval:period:day")],
        [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
        [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
    ]

    await q.edit_message_text(
        f"📊 **تقييم أداء المترجمين**\n\n"
        f"📅 الشهر: **{month_label} {year}**\n\n"
        f"اختر نوع الفترة:",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode=ParseMode.MARKDOWN,
    )
    return EVAL_SELECT_PERIOD


async def handle_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    if q.data == "eval:cancel":
        await q.edit_message_text("✅ تم إلغاء التقييم.")
        return ConversationHandler.END

    if q.data == "eval:back_month":
        year = context.user_data.get('eval_data', {}).get('year', date.today().year)
        keyboard = []
        for i in range(0, 12, 3):
            row = []
            for j in range(3):
                m = i + j + 1
                row.append(InlineKeyboardButton(
                    MONTH_NAMES[m], callback_data=f"eval:month:{m}"
                ))
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton("📄 كل الشهور", callback_data="eval:month:all")])
        keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_year")])
        keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n📅 السنة: **{year}**\n\nاختر الشهر:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_MONTH

    eval_data = context.user_data.setdefault('eval_data', {})
    year = eval_data.get('year', date.today().year)
    month = eval_data.get('month')
    if not month or month == "all":
        return EVAL_SELECT_MONTH

    if q.data == "eval:period:full":
        eval_data['period_type'] = "full"
        month_label = MONTH_NAMES.get(int(month), month)
        keyboard = [
            [InlineKeyboardButton("📄 PDF", callback_data="eval:format:pdf")],
            [InlineKeyboardButton("📊 Excel", callback_data="eval:format:excel")],
            [InlineKeyboardButton("📄 PDF + 📊 Excel", callback_data="eval:format:both")],
            [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
        ]
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n"
            f"📅 الفترة: **{month_label} {year}**\n\n"
            f"اختر صيغة الملف:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_FORMAT

    if q.data == "eval:period:day":
        eval_data['period_type'] = "day"
        month_int = int(month)
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n"
            f"📅 اختر يوماً من شهر **{MONTH_NAMES.get(month_int)} {year}**:",
            reply_markup=_build_day_calendar(year, month_int),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_DAY

    return EVAL_SELECT_PERIOD


async def handle_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()

    if q.data == "eval:cancel":
        await q.edit_message_text("✅ تم إلغاء التقييم.")
        return ConversationHandler.END

    if q.data == "eval:back_period":
        eval_data = context.user_data.get('eval_data', {})
        year = eval_data.get('year', date.today().year)
        month = eval_data.get('month')
        if not month or month == "all":
            return EVAL_SELECT_MONTH
        month_label = MONTH_NAMES.get(int(month), month)
        keyboard = [
            [InlineKeyboardButton("📅 الشهر كامل", callback_data="eval:period:full")],
            [InlineKeyboardButton("📆 يوم محدد", callback_data="eval:period:day")],
            [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
        ]
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n"
            f"📅 الشهر: **{month_label} {year}**\n\n"
            f"اختر نوع الفترة:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_PERIOD

    if q.data.startswith("evalday:select:"):
        date_str = q.data.split(":")[-1]
        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        eval_data = context.user_data.setdefault('eval_data', {})
        month = eval_data.get('month')
        year = eval_data.get('year')
        if month and year:
            if selected_date.month != int(month) or selected_date.year != int(year):
                await q.answer("اختر يوماً من نفس الشهر", show_alert=True)
                await q.edit_message_text(
                    f"📊 **تقييم أداء المترجمين**\n\n"
                    f"📅 اختر يوماً من شهر **{MONTH_NAMES.get(int(month))} {year}**:",
                    reply_markup=_build_day_calendar(int(year), int(month)),
                    parse_mode=ParseMode.MARKDOWN,
                )
                return EVAL_SELECT_DAY

        eval_data['day'] = selected_date
        period_label = format_date_arabic(selected_date)
        keyboard = [
            [InlineKeyboardButton("📄 PDF", callback_data="eval:format:pdf")],
            [InlineKeyboardButton("📊 Excel", callback_data="eval:format:excel")],
            [InlineKeyboardButton("📄 PDF + 📊 Excel", callback_data="eval:format:both")],
            [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
            [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
        ]
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n"
            f"📅 الفترة: **{period_label}**\n\n"
            f"اختر صيغة الملف:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_FORMAT

    return EVAL_SELECT_DAY


async def handle_format(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """اختيار التنسيق وتوليد التقرير"""
    q = update.callback_query
    await q.answer()

    if q.data == "eval:cancel":
        await q.edit_message_text("✅ تم إلغاء التقييم.")
        return ConversationHandler.END

    if q.data == "eval:back_month":
        eval_data = context.user_data.get('eval_data', {})
        year = eval_data.get('year', date.today().year)
        month = eval_data.get('month')
        period_type = eval_data.get('period_type')
        if month and month != "all" and period_type in ("full", "day"):
            month_label = MONTH_NAMES.get(int(month), month)
            keyboard = [
                [InlineKeyboardButton("📅 الشهر كامل", callback_data="eval:period:full")],
                [InlineKeyboardButton("📆 يوم محدد", callback_data="eval:period:day")],
                [InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_month")],
                [InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")],
            ]
            await q.edit_message_text(
                f"📊 **تقييم أداء المترجمين**\n\n"
                f"📅 الشهر: **{month_label} {year}**\n\n"
                f"اختر نوع الفترة:",
                reply_markup=InlineKeyboardMarkup(keyboard),
                parse_mode=ParseMode.MARKDOWN,
            )
            return EVAL_SELECT_PERIOD
        keyboard = []
        for i in range(0, 12, 3):
            row = []
            for j in range(3):
                m = i + j + 1
                row.append(InlineKeyboardButton(
                    MONTH_NAMES[m], callback_data=f"eval:month:{m}"
                ))
            keyboard.append(row)
        keyboard.append([InlineKeyboardButton("📄 كل الشهور", callback_data="eval:month:all")])
        keyboard.append([InlineKeyboardButton("🔙 رجوع", callback_data="eval:back_year")])
        keyboard.append([InlineKeyboardButton("❌ إلغاء", callback_data="eval:cancel")])
        await q.edit_message_text(
            f"📊 **تقييم أداء المترجمين**\n\n📅 السنة: **{year}**\n\nاختر الشهر:",
            reply_markup=InlineKeyboardMarkup(keyboard),
            parse_mode=ParseMode.MARKDOWN,
        )
        return EVAL_SELECT_MONTH

    fmt = q.data.split(":")[2]  # pdf, excel, both
    data = context.user_data.get('eval_data', {})
    year = data.get('year', date.today().year)
    month = data.get('month', 'all')
    period_type = data.get('period_type', 'full')

    if period_type == "day" and data.get('day'):
        day_date = data.get('day')
        period_label = format_date_arabic(day_date)
        start_date = day_date
        end_date = day_date
        start_date_str = day_date.strftime("%d/%m/%Y")
        end_date_str = day_date.strftime("%d/%m/%Y")
    elif period_type == "custom" and data.get('start_date') and data.get('end_date'):
        start_date = data.get('start_date')
        end_date = data.get('end_date')
        period_label = f"من {start_date} إلى {end_date}"
        start_date_str = start_date.strftime("%d/%m/%Y")
        end_date_str = end_date.strftime("%d/%m/%Y")
    else:
        month_label = "كل الشهور" if month == "all" else MONTH_NAMES.get(int(month), month)
        period_label = f"{month_label} {year}"
        start_date = None
        end_date = None
        start_date_str = None
        end_date_str = None

    await q.edit_message_text(
        f"⏳ **جاري إعداد تقرير التقييم...**\n\n"
        f"📅 الفترة: {period_label}\n"
        f"📄 الصيغة: {'PDF' if fmt == 'pdf' else 'Excel' if fmt == 'excel' else 'PDF + Excel'}",
        parse_mode=ParseMode.MARKDOWN,
    )

    try:
        with SessionLocal() as session:
            logger.info(f"📊 EVAL: period_type={period_type}, start={start_date}, end={end_date}, year={year}, month={month}")
            if (period_type == "day" or period_type == "custom") and start_date and end_date:
                raw_stats = get_translator_stats(session, start_date, end_date)
            else:
                raw_stats = get_monthly_stats(session, year, month)

            logger.info(f"📊 EVAL: raw_stats returned {len(raw_stats)} translators, total_reports={sum(r['total_reports'] for r in raw_stats)}")
            for i, rs in enumerate(raw_stats):
                logger.info(f"   ├ [{i+1}] tid={rs['translator_id']}, name={rs['translator_name']}, reports={rs['total_reports']}")

            if not raw_stats:
                await q.edit_message_text(
                    f"⚠️ **لا توجد تقارير في الفترة:** {period_label}\n\n"
                    "لا يمكن إنشاء تقييم بدون تقارير.",
                    parse_mode=ParseMode.MARKDOWN,
                )
                return ConversationHandler.END

            # تصفية لمترجم فردي إن اختير
            target_tid = data.get('target_translator_id')
            target_tname = (data.get('target_translator_name') or "").strip()
            if target_tname or target_tid:
                def _match(r):
                    if target_tid and r.get('translator_id') == target_tid:
                        return True
                    rn = _normalize_person_name(r.get('translator_name') or "")
                    tn = _normalize_person_name(target_tname)
                    return bool(rn and tn and rn == tn)
                raw_stats = [r for r in raw_stats if _match(r)]
                if not raw_stats:
                    await q.edit_message_text(
                        f"⚠️ **لا توجد تقارير للمترجم:** {target_tname}\n"
                        f"📅 خلال الفترة: {period_label}",
                        parse_mode=ParseMode.MARKDOWN,
                    )
                    return ConversationHandler.END

            # إضافة التقييم فوق الإحصائيات
            results = _compute_rating(raw_stats)

            # حفظ في قاعدة البيانات (فقط للشهور الكاملة أو السنة الكاملة، ولتقييم الكل فقط)
            if period_type not in ("day", "custom") and not (target_tname or target_tid):
                _save_evaluations_to_db(session, results, year, month)

            # لا نرسل ملخص/تفاصيل نصية على الشاشة — الإخراج يكون ملفات فقط

            # توليد وإرسال الملفات
            if target_tname:
                safe_name = re.sub(r"[^\w\-]+", "_", target_tname, flags=re.UNICODE).strip("_") or "مترجم"
                file_prefix = f"تقييم_فردي_{safe_name}_{year}"
                caption_pdf = f"📄 تقرير تقييم المترجم: {target_tname} - {period_label}"
                caption_xlsx = f"📊 تقرير Excel: {target_tname} - {period_label}"
            else:
                file_prefix = f"تقييم_المترجمين_{year}"
                caption_pdf = f"📄 تقرير تقييم المترجمين - {period_label}"
                caption_xlsx = f"📊 تقرير Excel - {period_label}"
            if period_type == "day" and data.get('day'):
                file_prefix += f"_{data['day'].strftime('%Y_%m_%d')}"
            elif period_type == "custom" and data.get('start_date') and data.get('end_date'):
                file_prefix += f"_{data['start_date'].strftime('%Y_%m_%d')}_إلى_{data['end_date'].strftime('%Y_%m_%d')}"
            elif month != "all":
                file_prefix += f"_{month}"

            if fmt in ('pdf', 'both'):
                try:
                    file_bytes, file_ext = _generate_pdf(results, period_label, year, month, start_date_str, end_date_str, target_name=target_tname or None)
                    file_obj = io.BytesIO(file_bytes)
                    file_obj.name = f"{file_prefix}.{file_ext}"
                    await q.message.reply_document(
                        document=file_obj,
                        caption=caption_pdf,
                    )
                    if file_ext != "pdf":
                        await q.message.reply_text("⚠️ تم إرسال HTML لأن PDF غير متوفر.")
                except Exception as e:
                    logger.error(f"خطأ في PDF: {e}", exc_info=True)
                    await q.message.reply_text(f"⚠️ خطأ في إنشاء PDF: {str(e)[:200]}")

            if fmt in ('excel', 'both'):
                try:
                    excel_bytes = _generate_excel(results, period_label, year, month, target_name=target_tname or None)
                    excel_file = io.BytesIO(excel_bytes)
                    excel_file.name = f"{file_prefix}.xlsx"
                    await q.message.reply_document(
                        document=excel_file,
                        caption=caption_xlsx,
                    )
                except Exception as e:
                    logger.error(f"خطأ في Excel: {e}", exc_info=True)
                    await q.message.reply_text(f"⚠️ خطأ في إنشاء Excel: {str(e)[:200]}")

    except Exception as e:
        logger.error(f"خطأ في التقييم: {e}", exc_info=True)
        # لا تستخدم Markdown هنا لأن نص الاستثناء قد يحتوي رموزًا تكسر parse entities.
        safe_error = str(e).replace("\n", " ").strip()[:300]
        await q.message.reply_text(f"❌ حدث خطأ: {safe_error}")

    return ConversationHandler.END


def _save_evaluations_to_db(session, results, year, month):
    """حفظ نتائج التقييم في قاعدة البيانات"""
    month_int = 0 if month == "all" else int(month)
    for res in results:
        try:
            existing = session.query(MonthlyEvaluation).filter_by(
                translator_name=res['translator_name'],
                year=year,
                month=month_int,
            ).first()

            if existing:
                existing.total_reports = res['total_reports']
                existing.work_days = res['work_days']
                existing.late_reports = res['late_reports']
                existing.total_points = res['final_score']
                existing.final_rating = int(res['final_score'] / 20)
                existing.performance_level = res['level']
                existing.updated_at = datetime.utcnow()
            else:
                ev = MonthlyEvaluation(
                    translator_id=res.get('translator_id'),
                    translator_name=res['translator_name'],
                    year=year,
                    month=month_int,
                    total_reports=res['total_reports'],
                    work_days=res['work_days'],
                    late_reports=res['late_reports'],
                    total_points=res['final_score'],
                    final_rating=int(res['final_score'] / 20),
                    performance_level=res['level'],
                )
                session.add(ev)
        except Exception as e:
            logger.warning(f"خطأ في حفظ تقييم {res['translator_name']}: {e}")
    try:
        session.commit()
    except Exception as e:
        logger.error(f"خطأ في حفظ التقييمات: {e}")
        session.rollback()


# ════════════════════════════════════════
# تسجيل الهاندلرز
# ════════════════════════════════════════

async def _cancel_evaluation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """إلغاء التقييم من زر inline"""
    query = update.callback_query
    if query:
        try:
            await query.answer()
        except Exception:
            pass
        try:
            await query.edit_message_text("✅ تم إلغاء التقييم.")
        except Exception:
            pass
    context.user_data.pop('eval_data', None)
    return ConversationHandler.END


async def _cancel_evaluation_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """إلغاء التقييم من أمر /cancel"""
    context.user_data.pop('eval_data', None)
    if update.message:
        await update.message.reply_text("✅ تم إلغاء التقييم.")
    return ConversationHandler.END


async def start_evaluation_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """نقطة دخول من زر inline (admin:evaluation)"""
    query = update.callback_query
    if query:
        await query.answer()

    user = update.effective_user
    if not is_admin(user.id):
        if query:
            await query.edit_message_text("هذه الخاصية مخصصة للأدمن فقط.")
        return ConversationHandler.END

    context.user_data.pop('eval_data', None)

    if query:
        await query.edit_message_text(
            _eval_intro_text(),
            reply_markup=_build_mode_keyboard(),
            parse_mode=ParseMode.MARKDOWN,
        )
    return EVAL_SELECT_MODE


def register(app):
    """تسجيل نظام تقييم المترجمين"""
    conv = ConversationHandler(
        entry_points=[
            MessageHandler(filters.Regex("^📊 تقييم المترجمين$"), start_evaluation),
            CallbackQueryHandler(start_evaluation_callback, pattern=r"^(admin:evaluation|eval_translators|translator_evaluation)$"),
        ],
        states={
            EVAL_SELECT_MODE: [
                CallbackQueryHandler(handle_eval_mode, pattern=r"^evalmode:"),
                CallbackQueryHandler(_cancel_evaluation, pattern=r"^eval:cancel$"),
            ],
            EVAL_SELECT_TRANSLATOR: [
                CallbackQueryHandler(handle_translator_pick, pattern=r"^evaltr:"),
                CallbackQueryHandler(handle_eval_mode, pattern=r"^evalmode:"),
                CallbackQueryHandler(_cancel_evaluation, pattern=r"^eval:cancel$"),
            ],
            EVAL_SELECT_YEAR: [
                CallbackQueryHandler(handle_eval_mode, pattern=r"^evalmode:back$"),
                CallbackQueryHandler(handle_year, pattern=r"^eval:"),
            ],
            EVAL_SELECT_MONTH: [
                CallbackQueryHandler(handle_month, pattern=r"^eval:"),
            ],
            EVAL_SELECT_PERIOD: [
                CallbackQueryHandler(handle_period, pattern=r"^eval:"),
            ],
            EVAL_SELECT_DAY: [
                CallbackQueryHandler(handle_day, pattern=r"^(evalday:|eval:)"),
            ],
            EVAL_SELECT_FORMAT: [
                CallbackQueryHandler(handle_format, pattern=r"^eval:"),
            ],
            EVAL_CUSTOM_START: [
                CallbackQueryHandler(handle_custom_date, pattern=r"^evalcustom:"),
                CallbackQueryHandler(_cancel_evaluation, pattern=r"^eval:cancel$"),
                CallbackQueryHandler(handle_month, pattern=r"^eval:back_year$"),
            ],
            EVAL_CUSTOM_END: [
                CallbackQueryHandler(handle_custom_date, pattern=r"^evalcustom:"),
                CallbackQueryHandler(handle_month, pattern=r"^eval:month:custom$"),
                CallbackQueryHandler(_cancel_evaluation, pattern=r"^eval:cancel$"),
            ],
        },
        fallbacks=[
            CommandHandler("cancel", _cancel_evaluation_command),
            CallbackQueryHandler(_cancel_evaluation, pattern=r"^eval:cancel$"),
            MessageHandler(filters.Regex("^📊 تقييم المترجمين$"), start_evaluation),
            CallbackQueryHandler(start_evaluation_callback, pattern=r"^(admin:evaluation|eval_translators|translator_evaluation)$"),
        ],
        name="translator_evaluation_conv",
        per_chat=True,
        per_user=True,
        per_message=False,
        allow_reentry=True,
    )
    app.add_handler(conv)
    logger.info("تم تسجيل نظام تقييم المترجمين")
